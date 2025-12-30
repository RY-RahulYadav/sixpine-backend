from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth import get_user_model
from django.db.models import Sum, Count, Q, Avg, F, DecimalField
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
import os
from urllib.parse import urlparse, unquote

from .permissions import IsAdminUser, IsAdminOrReadOnly
from .models import GlobalSettings, HomePageContent, BulkOrderPageContent, FAQPageContent, Advertisement
from .serializers import (
    DashboardStatsSerializer, AdminUserListSerializer, AdminUserDetailSerializer,
    AdminUserCreateSerializer, AdminUserUpdateSerializer, AdminCategorySerializer,
    AdminSubcategorySerializer, AdminColorSerializer, AdminMaterialSerializer,
    AdminProductListSerializer, AdminProductDetailSerializer,
    AdminOrderListSerializer, AdminOrderDetailSerializer, AdminDiscountSerializer,
    PaymentChargeSerializer, GlobalSettingsSerializer,
    AdminContactQuerySerializer, AdminBulkOrderSerializer, AdminLogSerializer,
    AdminCouponSerializer, HomePageContentSerializer, BulkOrderPageContentSerializer, FAQPageContentSerializer, AdvertisementSerializer,
    AdminDataRequestSerializer, AdminBrandSerializer, AdminBrandDetailSerializer,
    SellerOrderListSerializer, AdminMediaSerializer, AdminPackagingFeedbackSerializer,
    CategorySpecificationTemplateSerializer
)
from accounts.models import User, ContactQuery, BulkOrder, DataRequest, Vendor, Media, PackagingFeedback
from accounts.data_export_utils import export_orders_to_excel, export_addresses_to_excel, export_payment_options_to_excel
from products.excel_utils import generate_product_template, export_product_to_excel
from products.models import (
    Category, Subcategory, Color, Material, Product, ProductImage,
    ProductVariant, ProductVariantImage, ProductSpecification, ProductFeature,
    CategorySpecificationTemplate,
    ProductOffer, Discount, Coupon, ProductReview
)
from orders.models import Order, OrderItem, OrderStatusHistory, OrderNote
from .models import AdminLog
from .utils import create_admin_log
from .mixins import AdminLoggingMixin

User = get_user_model()


# ==================== Dashboard Views ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def dashboard_stats(request):
    """Get comprehensive dashboard statistics"""
    # Calculate date ranges
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)
    
    # Basic stats
    total_users = User.objects.count()
    total_orders = Order.objects.count()
    
    # Total order value = sum of total_amount (including tax) that customers paid
    # This is the actual amount customers paid, not just the subtotal
    total_order_value = Order.objects.aggregate(
        total=Sum(F('total_amount'), output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total'] or Decimal('0.00')
    
    # Separate calculations for seller orders and Sixpine orders
    delivered_orders = Order.objects.filter(status='delivered')
    
    # Net profit from sellers = tax + platform fee from delivered orders (excluding Sixpine products)
    # Get orders that contain seller products (not Sixpine)
    seller_delivered_orders = delivered_orders.exclude(
        Q(items__product__brand__iexact='Sixpine') | Q(items__product__vendor__isnull=True)
    ).distinct()
    
    seller_net_profit = seller_delivered_orders.aggregate(
        total=Sum(F('tax_amount') + F('platform_fee'), output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total'] or Decimal('0.00')
    
    # Sixpine profit = total order value of Sixpine products (full amount for delivered orders)
    # Get orders that contain Sixpine products
    sixpine_delivered_orders = delivered_orders.filter(
        Q(items__product__brand__iexact='Sixpine') | Q(items__product__vendor__isnull=True)
    ).distinct()
    
    sixpine_profit = Decimal('0.00')
    for order in sixpine_delivered_orders:
        # Get Sixpine items in this order
        sixpine_items = order.items.filter(
            Q(product__brand__iexact='Sixpine') | Q(product__vendor__isnull=True)
        )
        if sixpine_items.exists():
            # Calculate Sixpine's share of total amount (proportional to their items)
            sixpine_items_subtotal = sixpine_items.aggregate(
                total=Sum(F('price') * F('quantity'), output_field=DecimalField(max_digits=10, decimal_places=2))
            )['total'] or Decimal('0.00')
            
            if order.subtotal > 0:
                sixpine_share_ratio = sixpine_items_subtotal / order.subtotal
                sixpine_share_of_total = order.total_amount * sixpine_share_ratio
                sixpine_profit += sixpine_share_of_total
            else:
                sixpine_profit += sixpine_items_subtotal
    
    # Total net revenue = seller net profit + Sixpine profit
    total_net_revenue = seller_net_profit + sixpine_profit
    
    # Keep total_net_profit for backward compatibility (same as seller_net_profit)
    total_net_profit = seller_net_profit
    
    total_products = Product.objects.count()
    
    # Order summary stats
    orders_placed_count = Order.objects.count()
    delivered_orders_count = Order.objects.filter(status='delivered').count()
    cod_orders_count = Order.objects.filter(payment_method='COD').count()
    # Online payment = total orders - COD orders
    online_payment_orders_count = orders_placed_count - cod_orders_count
    
    # Low stock products - get global threshold, default to 100
    low_stock_threshold = GlobalSettings.get_setting('low_stock_threshold', 100)
    
    # Calculate low stock products: sum all variant stocks per product, compare to threshold
    products_with_stock = Product.objects.annotate(
        total_stock=Sum('variants__stock_quantity', filter=Q(variants__is_active=True))
    ).filter(total_stock__lt=low_stock_threshold, total_stock__isnull=False, is_active=True)
    low_stock_products = products_with_stock.count()
    
    # Recent orders (last 10)
    recent_orders = Order.objects.order_by('-created_at')[:10].values(
        'id', 'order_id', 'status', 'total_amount', 'created_at'
    )
    recent_orders = [{
        **order,
        'customer_name': Order.objects.get(id=order['id']).user.get_full_name() or Order.objects.get(id=order['id']).user.username
    } for order in recent_orders]
    
    # Top selling products with revenue calculation
    top_products = OrderItem.objects.values('product').annotate(
        sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('price'), output_field=DecimalField(max_digits=10, decimal_places=2))
    ).order_by('-sold')[:10]
    top_selling_products = []
    for item in top_products:
        product = Product.objects.get(id=item['product'])
        # Calculate revenue: sum of (quantity * price) for all order items of this product
        revenue = item.get('revenue') or Decimal('0.00')
        top_selling_products.append({
            'id': product.id,
            'title': product.title,
            'sold': item['sold'],
            'revenue': float(revenue)
        })
    
    # Sales by day (last 30 days)
    sales_by_day = []
    for i in range(30):
        date = thirty_days_ago + timedelta(days=i)
        day_orders = Order.objects.filter(
            created_at__date=date,
            payment_status='paid'
        )
        revenue = day_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        orders_count = day_orders.count()
        sales_by_day.append({
            'date': date.isoformat(),
            'revenue': float(revenue),
            'orders': orders_count
        })
    
    data = {
        'total_users': total_users,
        'total_orders': total_orders,
        'total_order_value': str(total_order_value),
        'total_net_profit': str(total_net_profit),  # Seller net profit (tax + platform fee)
        'seller_net_profit': str(seller_net_profit),  # Explicit seller net profit
        'sixpine_profit': str(sixpine_profit),  # Sixpine product profit (full order value)
        'total_net_revenue': str(total_net_revenue),  # Total net revenue (seller + Sixpine)
        'total_products': total_products,
        'orders_placed_count': orders_placed_count,
        'delivered_orders_count': delivered_orders_count,
        'cod_orders_count': cod_orders_count,
        'online_payment_orders_count': online_payment_orders_count,
        'low_stock_products': low_stock_products,
        'recent_orders': list(recent_orders),
        'top_selling_products': top_selling_products,
        'sales_by_day': sales_by_day
    }
    
    serializer = DashboardStatsSerializer(data)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def platform_analytics(request):
    """Get comprehensive platform analytics including seller profit, Sixpine products, and orders"""
    from decimal import Decimal
    from django.db.models import Count, Sum, F, Q, DecimalField
    from datetime import datetime, timedelta
    from accounts.models import Vendor
    
    # Calculate date ranges
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)
    twelve_months_ago = today - timedelta(days=365)
    
    # ========== ORDER STATS ==========
    total_orders = Order.objects.count()
    delivered_orders = Order.objects.filter(status='delivered')
    
    # Total order value (all orders)
    total_order_value = Order.objects.aggregate(
        total=Sum(F('total_amount'), output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total'] or Decimal('0.00')
    
    # Orders by status
    orders_by_status = Order.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Orders by month (last 12 months)
    orders_by_month = []
    for i in range(12):
        month_start = today.replace(day=1) - timedelta(days=30 * i)
        month_orders = Order.objects.filter(
            created_at__year=month_start.year,
            created_at__month=month_start.month
        )
        month_order_value = month_orders.aggregate(
            total=Sum(F('total_amount'), output_field=DecimalField(max_digits=10, decimal_places=2))
        )['total'] or Decimal('0.00')
        
        orders_by_month.append({
            'month': f"{month_start.year}-{str(month_start.month).zfill(2)}",
            'count': month_orders.count(),
            'order_value': float(month_order_value)
        })
    orders_by_month.reverse()
    
    # Payment methods
    payment_methods = Order.objects.values('payment_method').annotate(
        count=Count('id'),
        total_value=Sum(F('total_amount'), output_field=DecimalField(max_digits=10, decimal_places=2))
    ).order_by('-count')
    
    # ========== SELLER PROFIT STATS ==========
    # Get orders with seller products (not Sixpine)
    seller_orders = Order.objects.filter(
        items__product__vendor__isnull=False
    ).exclude(
        Q(items__product__brand__iexact='Sixpine')
    ).distinct()
    
    seller_delivered_orders = seller_orders.filter(status='delivered')
    
    # Seller net profit = tax + platform fee from delivered orders
    seller_net_profit = seller_delivered_orders.aggregate(
        total=Sum(F('tax_amount') + F('platform_fee'), output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total'] or Decimal('0.00')
    
    # Seller order value (proportional share of total_amount)
    seller_order_value = Decimal('0.00')
    for order in seller_orders:
        seller_items = order.items.filter(product__vendor__isnull=False).exclude(
            Q(product__brand__iexact='Sixpine')
        )
        if seller_items.exists():
            seller_items_subtotal = seller_items.aggregate(
                total=Sum(F('price') * F('quantity'), output_field=DecimalField(max_digits=10, decimal_places=2))
            )['total'] or Decimal('0.00')
            
            if order.subtotal > 0:
                seller_share_ratio = seller_items_subtotal / order.subtotal
                seller_share_of_total = order.total_amount * seller_share_ratio
                seller_order_value += seller_share_of_total
            else:
                seller_order_value += seller_items_subtotal
    
    # Seller orders by month
    seller_orders_by_month = []
    for i in range(12):
        month_start = today.replace(day=1) - timedelta(days=30 * i)
        month_seller_orders = seller_orders.filter(
            created_at__year=month_start.year,
            created_at__month=month_start.month
        )
        
        month_seller_value = Decimal('0.00')
        for order in month_seller_orders:
            seller_items = order.items.filter(product__vendor__isnull=False).exclude(
                Q(product__brand__iexact='Sixpine')
            )
            if seller_items.exists():
                seller_items_subtotal = seller_items.aggregate(
                    total=Sum(F('price') * F('quantity'), output_field=DecimalField(max_digits=10, decimal_places=2))
                )['total'] or Decimal('0.00')
                
                if order.subtotal > 0:
                    seller_share_ratio = seller_items_subtotal / order.subtotal
                    seller_share_of_total = order.total_amount * seller_share_ratio
                    month_seller_value += seller_share_of_total
                else:
                    month_seller_value += seller_items_subtotal
        
        seller_orders_by_month.append({
            'month': f"{month_start.year}-{str(month_start.month).zfill(2)}",
            'count': month_seller_orders.count(),
            'order_value': float(month_seller_value),
            'net_profit': float(seller_delivered_orders.filter(
                created_at__year=month_start.year,
                created_at__month=month_start.month
            ).aggregate(
                total=Sum(F('tax_amount') + F('platform_fee'), output_field=DecimalField(max_digits=10, decimal_places=2))
            )['total'] or Decimal('0.00'))
        })
    seller_orders_by_month.reverse()
    
    # ========== SIXPINE PRODUCTS STATS ==========
    # Get Sixpine products (brand='Sixpine' or vendor is null)
    sixpine_products = Product.objects.filter(
        Q(brand__iexact='Sixpine') | Q(vendor__isnull=True)
    )
    total_sixpine_products = sixpine_products.count()
    active_sixpine_products = sixpine_products.filter(is_active=True).count()
    
    # Get orders with Sixpine products
    sixpine_orders = Order.objects.filter(
        Q(items__product__brand__iexact='Sixpine') | Q(items__product__vendor__isnull=True)
    ).distinct()
    
    sixpine_delivered_orders = sixpine_orders.filter(status='delivered')
    
    # Sixpine profit = total order value of Sixpine products (full amount for delivered orders)
    sixpine_profit = Decimal('0.00')
    for order in sixpine_delivered_orders:
        sixpine_items = order.items.filter(
            Q(product__brand__iexact='Sixpine') | Q(product__vendor__isnull=True)
        )
        if sixpine_items.exists():
            sixpine_items_subtotal = sixpine_items.aggregate(
                total=Sum(F('price') * F('quantity'), output_field=DecimalField(max_digits=10, decimal_places=2))
            )['total'] or Decimal('0.00')
            
            if order.subtotal > 0:
                sixpine_share_ratio = sixpine_items_subtotal / order.subtotal
                sixpine_share_of_total = order.total_amount * sixpine_share_ratio
                sixpine_profit += sixpine_share_of_total
            else:
                sixpine_profit += sixpine_items_subtotal
    
    # Sixpine order value (all orders, not just delivered)
    sixpine_order_value = Decimal('0.00')
    for order in sixpine_orders:
        sixpine_items = order.items.filter(
            Q(product__brand__iexact='Sixpine') | Q(product__vendor__isnull=True)
        )
        if sixpine_items.exists():
            sixpine_items_subtotal = sixpine_items.aggregate(
                total=Sum(F('price') * F('quantity'), output_field=DecimalField(max_digits=10, decimal_places=2))
            )['total'] or Decimal('0.00')
            
            if order.subtotal > 0:
                sixpine_share_ratio = sixpine_items_subtotal / order.subtotal
                sixpine_share_of_total = order.total_amount * sixpine_share_ratio
                sixpine_order_value += sixpine_share_of_total
            else:
                sixpine_order_value += sixpine_items_subtotal
    
    # Sixpine orders by month
    sixpine_orders_by_month = []
    for i in range(12):
        month_start = today.replace(day=1) - timedelta(days=30 * i)
        month_sixpine_orders = sixpine_orders.filter(
            created_at__year=month_start.year,
            created_at__month=month_start.month
        )
        
        month_sixpine_value = Decimal('0.00')
        for order in month_sixpine_orders:
            sixpine_items = order.items.filter(
                Q(product__brand__iexact='Sixpine') | Q(product__vendor__isnull=True)
            )
            if sixpine_items.exists():
                sixpine_items_subtotal = sixpine_items.aggregate(
                    total=Sum(F('price') * F('quantity'), output_field=DecimalField(max_digits=10, decimal_places=2))
                )['total'] or Decimal('0.00')
                
                if order.subtotal > 0:
                    sixpine_share_ratio = sixpine_items_subtotal / order.subtotal
                    sixpine_share_of_total = order.total_amount * sixpine_share_ratio
                    month_sixpine_value += sixpine_share_of_total
                else:
                    month_sixpine_value += sixpine_items_subtotal
        
        # Calculate Sixpine profit for this month (from delivered orders)
        month_sixpine_profit = Decimal('0.00')
        month_delivered_sixpine_orders = sixpine_delivered_orders.filter(
            created_at__year=month_start.year,
            created_at__month=month_start.month
        )
        for order in month_delivered_sixpine_orders:
            sixpine_items = order.items.filter(
                Q(product__brand__iexact='Sixpine') | Q(product__vendor__isnull=True)
            )
            if sixpine_items.exists():
                sixpine_items_subtotal = sixpine_items.aggregate(
                    total=Sum(F('price') * F('quantity'), output_field=DecimalField(max_digits=10, decimal_places=2))
                )['total'] or Decimal('0.00')
                
                if order.subtotal > 0:
                    sixpine_share_ratio = sixpine_items_subtotal / order.subtotal
                    sixpine_share_of_total = order.total_amount * sixpine_share_ratio
                    month_sixpine_profit += sixpine_share_of_total
                else:
                    month_sixpine_profit += sixpine_items_subtotal
        
        sixpine_orders_by_month.append({
            'month': f"{month_start.year}-{str(month_start.month).zfill(2)}",
            'count': month_sixpine_orders.count(),
            'order_value': float(month_sixpine_value),
            'profit': float(month_sixpine_profit)
        })
    sixpine_orders_by_month.reverse()
    
    # Top selling Sixpine products
    sixpine_top_products = OrderItem.objects.filter(
        Q(product__brand__iexact='Sixpine') | Q(product__vendor__isnull=True)
    ).values('product').annotate(
        sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('price'), output_field=DecimalField(max_digits=10, decimal_places=2))
    ).order_by('-sold')[:10]
    
    top_sixpine_products = []
    for item in sixpine_top_products:
        try:
            product = Product.objects.get(id=item['product'])
            top_sixpine_products.append({
                'id': product.id,
                'title': product.title,
                'sold': item['sold'],
                'revenue': float(item.get('revenue') or Decimal('0.00'))
            })
        except Product.DoesNotExist:
            continue
    
    # ========== PLATFORM SUMMARY ==========
    total_net_revenue = seller_net_profit + sixpine_profit
    total_products = Product.objects.count()
    total_seller_products = Product.objects.filter(vendor__isnull=False).exclude(
        Q(brand__iexact='Sixpine')
    ).count()
    
    # Total vendors
    total_vendors = Vendor.objects.filter(status='active').count()
    
    data = {
        'order_stats': {
            'total_orders': total_orders,
            'total_order_value': float(total_order_value),
            'average_order_value': float(total_order_value / total_orders) if total_orders > 0 else 0,
            'orders_by_status': [
                {'status': item['status'], 'count': item['count']}
                for item in orders_by_status
            ],
            'orders_by_month': orders_by_month,
            'payment_methods': [
                {
                    'method': item['payment_method'] or 'Unknown',
                    'count': item['count'],
                    'total_value': float(item.get('total_value') or Decimal('0.00'))
                }
                for item in payment_methods
            ]
        },
        'seller_stats': {
            'total_vendors': total_vendors,
            'total_seller_products': total_seller_products,
            'total_seller_orders': seller_orders.count(),
            'seller_order_value': float(seller_order_value),
            'seller_net_profit': float(seller_net_profit),
            'orders_by_month': seller_orders_by_month
        },
        'sixpine_stats': {
            'total_products': total_sixpine_products,
            'active_products': active_sixpine_products,
            'total_orders': sixpine_orders.count(),
            'sixpine_order_value': float(sixpine_order_value),
            'sixpine_profit': float(sixpine_profit),
            'orders_by_month': sixpine_orders_by_month,
            'top_products': top_sixpine_products
        },
        'platform_summary': {
            'total_products': total_products,
            'total_order_value': float(total_order_value),
            'total_net_revenue': float(total_net_revenue),
            'seller_net_profit': float(seller_net_profit),
            'sixpine_profit': float(sixpine_profit)
        }
    }
    
    return Response(data)


# ==================== User Management Views ====================
class AdminUserViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for user management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = AdminUserListSerializer
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return AdminUserDetailSerializer
        elif self.action == 'create':
            return AdminUserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return AdminUserUpdateSerializer
        return AdminUserListSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        is_active = self.request.query_params.get('is_active', None)
        is_staff = self.request.query_params.get('is_staff', None)
        
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(mobile__icontains=search)
            )
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        if is_staff is not None:
            queryset = queryset.filter(is_staff=is_staff.lower() == 'true')
        
        return queryset
    
    def perform_destroy(self, instance):
        """Override destroy to check for orders and handle ProtectedError"""
        from django.db.models.deletion import ProtectedError
        from rest_framework.exceptions import ValidationError, PermissionDenied
        
        # Check if user is superuser
        if instance.is_superuser:
            raise PermissionDenied("Cannot delete superuser accounts")
        
        # Check if user has orders
        order_count = Order.objects.filter(user=instance).count()
        if order_count > 0:
            raise ValidationError(
                f"Cannot delete user because they have {order_count} order(s). "
                "Please delete or reassign the orders first."
            )
        
        # Try to delete and catch ProtectedError
        try:
            # Log before deletion
            try:
                create_admin_log(
                    request=self.request,
                    action_type='delete',
                    model_name='User',
                    object_id=instance.id,
                    object_repr=str(instance),
                    details={'action': 'delete'}
                )
            except Exception as e:
                print(f"Error creating admin log: {e}")
            
            instance.delete()
        except ProtectedError as e:
            # Extract order information from the error
            protected_objects = list(e.protected_objects)
            order_count = len([obj for obj in protected_objects if isinstance(obj, Order)])
            
            if order_count > 0:
                raise ValidationError(
                    f"Cannot delete user because they have {order_count} order(s). "
                    "Please delete or reassign the orders first."
                )
            else:
                raise ValidationError(
                    "Cannot delete user because they are referenced by other records. "
                    "Please remove all associated data first."
                )
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Toggle user active status"""
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        
        create_admin_log(
            request=request,
            action_type='activate' if user.is_active else 'deactivate',
            model_name='User',
            object_id=user.id,
            object_repr=str(user)
        )
        
        serializer = self.get_serializer(user)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def toggle_staff(self, request, pk=None):
        """Toggle user staff status"""
        user = self.get_object()
        user.is_staff = not user.is_staff
        user.save()
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='User',
            object_id=user.id,
            object_repr=str(user),
            details={'is_staff': user.is_staff}
        )
        
        serializer = self.get_serializer(user)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Reset user password"""
        user = self.get_object()
        password = request.data.get('password')
        
        if not password:
            return Response(
                {'error': 'Password is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.set_password(password)
        user.save()
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='User',
            object_id=user.id,
            object_repr=str(user),
            details={'action': 'password_reset'}
        )
        
        return Response({'message': 'Password reset successfully'})


# ==================== Category Management Views ====================
class AdminCategoryViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for category management"""
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    queryset = Category.objects.all().order_by('sort_order', 'name')
    serializer_class = AdminCategorySerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def hierarchical(self, request):
        """Get categories with subcategories"""
        categories = Category.objects.filter(is_active=True).prefetch_related('subcategories')
        serializer = self.get_serializer(categories, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def specification_defaults(self, request, pk=None):
        """Get default specification field names for a category"""
        category = self.get_object()
        templates = CategorySpecificationTemplate.objects.filter(
            category=category,
            is_active=True
        ).order_by('section', 'sort_order', 'field_name')
        
        # Group by section
        defaults = {
            'specifications': [],
            'measurement_specs': [],
            'style_specs': [],
            'features': [],
            'user_guide': [],
            'item_details': []
        }
        
        for template in templates:
            defaults[template.section].append({
                'field_name': template.field_name,
                'sort_order': template.sort_order
            })
        
        return Response(defaults)
    
    @action(detail=True, methods=['get'])
    def download_excel_template(self, request, pk=None):
        """Download category-specific Excel template for bulk product creation"""
        from django.http import HttpResponse
        from io import BytesIO
        
        try:
            category = self.get_object()
            wb = generate_product_template(category.id)
            
            # Create in-memory file
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create HTTP response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="product_template_{category.slug}_{category.id}.xlsx"'
            
            # Log action
            create_admin_log(
                request=request,
                action_type='download',
                model_name='Category',
                object_id=category.id,
                object_repr=str(category),
                details={'action': 'download_excel_template'}
            )
            
            return response
        except Exception as e:
            return Response({
                'error': f'Failed to generate template: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== Category Specification Template Management ====================
class AdminCategorySpecificationTemplateViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for managing category specification templates"""
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    queryset = CategorySpecificationTemplate.objects.all().order_by('category', 'section', 'sort_order', 'field_name')
    serializer_class = CategorySpecificationTemplateSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        category = self.request.query_params.get('category', None)
        section = self.request.query_params.get('section', None)
        
        if category:
            queryset = queryset.filter(category_id=category)
        if section:
            queryset = queryset.filter(section=section)
        
        return queryset


# ==================== Subcategory Management Views ====================
class AdminSubcategoryViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for subcategory management"""
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    queryset = Subcategory.objects.all().order_by('sort_order', 'name')
    serializer_class = AdminSubcategorySerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        category = self.request.query_params.get('category', None)
        search = self.request.query_params.get('search', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if category:
            queryset = queryset.filter(category_id=category)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset


# ==================== Color Management Views ====================
class AdminColorViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for color management"""
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    queryset = Color.objects.all().order_by('name')
    serializer_class = AdminColorSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        return queryset


# ==================== Material Management Views ====================
class AdminMaterialViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for material management"""
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    queryset = Material.objects.all().order_by('name')
    serializer_class = AdminMaterialSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        return queryset


# ==================== Product Management Views ====================
class AdminProductViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for product management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = Product.objects.all().select_related('category', 'subcategory').prefetch_related(
        'images', 'variants', 'variants__specifications', 'variants__subcategories', 'features', 'about_items'
    ).order_by('-created_at')
    serializer_class = AdminProductListSerializer
    
    def get_serializer_class(self):
        if self.action in ['retrieve', 'create', 'update', 'partial_update']:
            return AdminProductDetailSerializer
        return AdminProductListSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        category = self.request.query_params.get('category', None)
        subcategory = self.request.query_params.get('subcategory', None)
        is_active = self.request.query_params.get('is_active', None)
        is_featured = self.request.query_params.get('is_featured', None)
        vendor = self.request.query_params.get('vendor', None)
        brand = self.request.query_params.get('brand', None)
        
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(sku__icontains=search) |
                Q(short_description__icontains=search) |
                Q(long_description__icontains=search) |
                Q(brand__icontains=search)
            )
        if category:
            queryset = queryset.filter(category_id=category)
        if subcategory:
            queryset = queryset.filter(subcategory_id=subcategory)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        if is_featured is not None:
            queryset = queryset.filter(is_featured=is_featured.lower() == 'true')
        if vendor:
            queryset = queryset.filter(vendor_id=vendor)
        if brand:
            queryset = queryset.filter(brand__iexact=brand)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Toggle product active status"""
        product = self.get_object()
        product.is_active = not product.is_active
        product.save()
        
        create_admin_log(
            request=request,
            action_type='activate' if product.is_active else 'deactivate',
            model_name='Product',
            object_id=product.id,
            object_repr=str(product)
        )
        
        serializer = self.get_serializer(product)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def toggle_featured(self, request, pk=None):
        """Toggle product featured status"""
        product = self.get_object()
        product.is_featured = not product.is_featured
        product.save()
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='Product',
            object_id=product.id,
            object_repr=str(product),
            details={'is_featured': product.is_featured}
        )
        
        serializer = self.get_serializer(product)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def update_stock(self, request, pk=None):
        """Update product variant stock"""
        product = self.get_object()
        variant_id = request.data.get('variant_id')
        quantity = request.data.get('quantity')
        
        if not variant_id or quantity is None:
            return Response(
                {'error': 'variant_id and quantity are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            variant = product.variants.get(id=variant_id)
            variant.stock_quantity = quantity
            variant.save()
            
            create_admin_log(
                request=request,
                action_type='update',
                model_name='ProductVariant',
                object_id=variant.id,
                object_repr=str(variant),
                details={'stock_quantity': quantity}
            )
            
            return Response({
                'message': 'Stock updated successfully',
                'variant_id': variant.id,
                'stock_quantity': variant.stock_quantity
            })
        except ProductVariant.DoesNotExist:
            return Response(
                {'error': 'Variant not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['get'])
    def download_excel(self, request, pk=None):
        """Download existing product data as Excel"""
        from django.http import HttpResponse
        from io import BytesIO
        
        try:
            product = self.get_object()
            wb = export_product_to_excel(product.id)
            
            # Create in-memory file
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create HTTP response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="product_{product.slug}_{product.id}.xlsx"'
            
            # Log action
            create_admin_log(
                request=request,
                action_type='download',
                model_name='Product',
                object_id=product.id,
                object_repr=str(product),
                details={'action': 'download_excel'}
            )
            
            return response
        except Exception as e:
            return Response({
                'error': f'Failed to export product: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def update_from_excel(self, request, pk=None):
        """Update existing product from Excel file with two-tab format (Parent Product and Child Variation)"""
        from openpyxl import load_workbook
        from io import BytesIO
        from django.utils.text import slugify
        from decimal import Decimal, InvalidOperation
        from products.models import (
            ProductSpecification, VariantMeasurementSpec, VariantStyleSpec,
            VariantFeature, VariantUserGuide, VariantItemDetail,
            ProductAboutItem, ProductVariantImage
        )
        
        product = self.get_object()
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Excel file is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        excel_file = request.FILES['file']
        
        try:
            # Load workbook
            wb = load_workbook(filename=BytesIO(excel_file.read()), data_only=True)
            
            # Check for required sheets
            if 'Parent Product' not in wb.sheetnames or 'Child Variation' not in wb.sheetnames:
                return Response(
                    {'error': 'Excel file must contain "Parent Product" and "Child Variation" sheets'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            ws_parent = wb['Parent Product']
            ws_child = wb['Child Variation']
            
            # Get headers from both sheets
            parent_headers = [cell.value for cell in ws_parent[1]]
            child_headers = [cell.value for cell in ws_child[1]]
            
            # Create column maps
            parent_col_map = {}
            for idx, header in enumerate(parent_headers, 1):
                if header:
                    parent_col_map[str(header).strip()] = idx
            
            child_col_map = {}
            for idx, header in enumerate(child_headers, 1):
                if header:
                    child_col_map[str(header).strip()] = idx
            
            # Validate required columns
            required_parent = ['SKU*']
            required_child = ['Parent SKU*', 'Variant Color*', 'Variant Price*', 'Variant Stock Quantity*']
            
            missing_parent = [f for f in required_parent if f not in parent_col_map]
            missing_child = [f for f in required_child if f not in child_col_map]
            
            if missing_parent or missing_child:
                errors_list = []
                if missing_parent:
                    errors_list.append(f'Parent Product tab missing: {", ".join(missing_parent)}')
                if missing_child:
                    errors_list.append(f'Child Variation tab missing: {", ".join(missing_child)}')
                return Response(
                    {'error': '; '.join(errors_list)},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            variants_created = 0
            variants_updated = 0
            errors = []
            product_updated = False
            
            # Step 1: Update product from Parent Product tab (find row with matching SKU)
            parent_row = None
            for row in ws_parent.iter_rows(min_row=2, values_only=False):
                if not row[parent_col_map['SKU*'] - 1].value:
                    continue
                
                sku = str(row[parent_col_map['SKU*'] - 1].value).strip()
                if sku == product.sku or (not product.sku and sku):
                    parent_row = row
                    break
            
            if parent_row:
                # Update product-level fields only if they have values
                if parent_col_map.get('Product Title*') and parent_row[parent_col_map['Product Title*'] - 1].value:
                    new_title = str(parent_row[parent_col_map['Product Title*'] - 1].value).strip()
                    if new_title and new_title != product.title:
                        product.title = new_title
                        product_updated = True
                
                if parent_col_map.get('Short Description*') and parent_row[parent_col_map['Short Description*'] - 1].value:
                    new_desc = str(parent_row[parent_col_map['Short Description*'] - 1].value).strip()
                    if new_desc != product.short_description:
                        product.short_description = new_desc
                        product_updated = True
                
                if parent_col_map.get('Long Description') and parent_row[parent_col_map['Long Description'] - 1].value:
                    new_long_desc = str(parent_row[parent_col_map['Long Description'] - 1].value).strip()
                    if new_long_desc != product.long_description:
                        product.long_description = new_long_desc
                        product_updated = True
                
                # Material
                if parent_col_map.get('Material') and parent_row[parent_col_map['Material'] - 1].value:
                    material_name = str(parent_row[parent_col_map['Material'] - 1].value).strip()
                    try:
                        material = Material.objects.get(name=material_name)
                        if product.material != material:
                            product.material = material
                            product_updated = True
                    except Material.DoesNotExist:
                        pass
                
                # Other fields
                if parent_col_map.get('Brand') and parent_row[parent_col_map['Brand'] - 1].value:
                    new_brand = str(parent_row[parent_col_map['Brand'] - 1].value).strip()
                    if new_brand != product.brand:
                        product.brand = new_brand
                        product_updated = True
                
                if parent_col_map.get('Dimensions') and parent_row[parent_col_map['Dimensions'] - 1].value:
                    new_dims = str(parent_row[parent_col_map['Dimensions'] - 1].value).strip()
                    if new_dims != product.dimensions:
                        product.dimensions = new_dims
                        product_updated = True
                
                if parent_col_map.get('Weight') and parent_row[parent_col_map['Weight'] - 1].value:
                    new_weight = str(parent_row[parent_col_map['Weight'] - 1].value).strip()
                    if new_weight != product.weight:
                        product.weight = new_weight
                        product_updated = True
                
                if parent_col_map.get('Warranty') and parent_row[parent_col_map['Warranty'] - 1].value:
                    new_warranty = str(parent_row[parent_col_map['Warranty'] - 1].value).strip()
                    if new_warranty != product.warranty:
                        product.warranty = new_warranty
                        product_updated = True
                
                if parent_col_map.get('Assembly Required') and parent_row[parent_col_map['Assembly Required'] - 1].value:
                    assembly_val = str(parent_row[parent_col_map['Assembly Required'] - 1].value).strip().lower()
                    new_assembly = assembly_val == 'yes'
                    if new_assembly != product.assembly_required:
                        product.assembly_required = new_assembly
                        product_updated = True
                
                if parent_col_map.get('Estimated Delivery Days') and parent_row[parent_col_map['Estimated Delivery Days'] - 1].value:
                    try:
                        new_delivery = int(parent_row[parent_col_map['Estimated Delivery Days'] - 1].value)
                        if new_delivery != product.estimated_delivery_days:
                            product.estimated_delivery_days = new_delivery
                            product_updated = True
                    except:
                        pass
                
                if parent_col_map.get('Care Instructions') and parent_row[parent_col_map['Care Instructions'] - 1].value:
                    new_care = str(parent_row[parent_col_map['Care Instructions'] - 1].value).strip()
                    if new_care != product.care_instructions:
                        product.care_instructions = new_care
                        product_updated = True
                
                if parent_col_map.get('What is in Box') and parent_row[parent_col_map['What is in Box'] - 1].value:
                    new_box = str(parent_row[parent_col_map['What is in Box'] - 1].value).strip()
                    if new_box != product.what_in_box:
                        product.what_in_box = new_box
                        product_updated = True
                
                if parent_col_map.get('Meta Title') and parent_row[parent_col_map['Meta Title'] - 1].value:
                    new_meta_title = str(parent_row[parent_col_map['Meta Title'] - 1].value).strip()
                    if new_meta_title != product.meta_title:
                        product.meta_title = new_meta_title
                        product_updated = True
                
                if parent_col_map.get('Meta Description') and parent_row[parent_col_map['Meta Description'] - 1].value:
                    new_meta_desc = str(parent_row[parent_col_map['Meta Description'] - 1].value).strip()
                    if new_meta_desc != product.meta_description:
                        product.meta_description = new_meta_desc
                        product_updated = True
                
                if parent_col_map.get('Is Featured') and parent_row[parent_col_map['Is Featured'] - 1].value:
                    featured_val = str(parent_row[parent_col_map['Is Featured'] - 1].value).strip().lower()
                    new_featured = featured_val == 'yes'
                    if new_featured != product.is_featured:
                        product.is_featured = new_featured
                        product_updated = True
                
                if parent_col_map.get('Is Active') and parent_row[parent_col_map['Is Active'] - 1].value:
                    active_val = str(parent_row[parent_col_map['Is Active'] - 1].value).strip().lower()
                    new_active = active_val == 'yes'
                    if new_active != product.is_active:
                        product.is_active = new_active
                        product_updated = True
                
                # Update About Items (clear and recreate)
                if any(parent_col_map.get(f'About This Item {i}') for i in range(1, 11)):
                    product.about_items.all().delete()
                    for i in range(1, 11):
                        col_name = f'About This Item {i}'
                        if parent_col_map.get(col_name) and parent_row[parent_col_map[col_name] - 1].value:
                            item_text = str(parent_row[parent_col_map[col_name] - 1].value).strip()
                            if item_text:
                                ProductAboutItem.objects.create(
                                    product=product,
                                    item=item_text,
                                    sort_order=i
                                )
                                product_updated = True
                
                # Update Screen Offers (clear and recreate)
                screen_offers = []
                for i in range(1, 11):
                    title_col = f'Screen Offer Title {i}'
                    desc_col = f'Screen Offer Description {i}'
                    if parent_col_map.get(title_col) and parent_row[parent_col_map[title_col] - 1].value:
                        offer_title = str(parent_row[parent_col_map[title_col] - 1].value).strip()
                        offer_desc = ''
                        if parent_col_map.get(desc_col) and parent_row[parent_col_map[desc_col] - 1].value:
                            offer_desc = str(parent_row[parent_col_map[desc_col] - 1].value).strip()
                        if offer_title:
                            screen_offers.append({'title': offer_title, 'description': offer_desc})
                
                if screen_offers != product.screen_offer:
                    product.screen_offer = screen_offers
                    product_updated = True
            
            if product_updated:
                product.save()
            
            # Step 2: Update/Create variants from Child Variation tab
            for row_num, row in enumerate(ws_child.iter_rows(min_row=2, values_only=False), start=2):
                # Skip empty rows
                if not row[child_col_map['Parent SKU*'] - 1].value:
                    continue
                
                try:
                    parent_sku = str(row[child_col_map['Parent SKU*'] - 1].value).strip()
                    if not parent_sku or parent_sku == 'Parent SKU*':
                        continue
                    
                    # Verify this variant belongs to this product
                    if parent_sku != product.sku:
                        continue  # Skip variants that don't belong to this product
                    
                    # Get color
                    color_name = str(row[child_col_map['Variant Color*'] - 1].value).strip()
                    try:
                        color = Color.objects.get(name=color_name)
                    except Color.DoesNotExist:
                        errors.append(f'Child row {row_num}: Color "{color_name}" not found')
                        continue
                    
                    # Variant fields
                    size = str(row[child_col_map.get('Variant Size', 0) - 1].value).strip() if child_col_map.get('Variant Size') and row[child_col_map.get('Variant Size', 0) - 1].value else ''
                    pattern = str(row[child_col_map.get('Variant Pattern', 0) - 1].value).strip() if child_col_map.get('Variant Pattern') and row[child_col_map.get('Variant Pattern', 0) - 1].value else ''
                    quality = str(row[child_col_map.get('Variant Quality', 0) - 1].value).strip() if child_col_map.get('Variant Quality') and row[child_col_map.get('Variant Quality', 0) - 1].value else ''
                    variant_title = str(row[child_col_map.get('Variant Title', 0) - 1].value).strip() if child_col_map.get('Variant Title') and row[child_col_map.get('Variant Title', 0) - 1].value else ''
                    
                    # Try to find existing variant
                    variant = product.variants.filter(
                        color=color,
                        size=size,
                        pattern=pattern,
                        quality=quality
                    ).first()
                    
                    # Price
                    try:
                        price = Decimal(str(row[child_col_map['Variant Price*'] - 1].value))
                    except (ValueError, InvalidOperation, TypeError):
                        errors.append(f'Child row {row_num}: Invalid price value')
                        continue
                    
                    old_price = None
                    if child_col_map.get('Variant Old Price') and row[child_col_map['Variant Old Price'] - 1].value:
                        try:
                            old_price = Decimal(str(row[child_col_map['Variant Old Price'] - 1].value))
                        except:
                            pass
                    
                    # Stock
                    try:
                        stock_qty = int(row[child_col_map['Variant Stock Quantity*'] - 1].value or 0)
                    except (ValueError, TypeError):
                        errors.append(f'Child row {row_num}: Invalid stock quantity')
                        continue
                    
                    is_in_stock = True
                    if child_col_map.get('Variant Is In Stock') and row[child_col_map['Variant Is In Stock'] - 1].value:
                        stock_val = str(row[child_col_map['Variant Is In Stock'] - 1].value).strip().lower()
                        is_in_stock = stock_val == 'yes'
                    
                    is_active = True
                    if child_col_map.get('Variant Is Active') and row[child_col_map['Variant Is Active'] - 1].value:
                        active_val = str(row[child_col_map['Variant Is Active'] - 1].value).strip().lower()
                        is_active = active_val == 'yes'
                    
                    # Variant image
                    variant_image = ''
                    if child_col_map.get('Variant Image URL') and row[child_col_map['Variant Image URL'] - 1].value:
                        variant_image = str(row[child_col_map['Variant Image URL'] - 1].value).strip()
                    
                    if variant:
                        # Update existing variant (only if changed)
                        updated = False
                        if variant.price != price:
                            variant.price = price
                            updated = True
                        if variant.old_price != old_price:
                            variant.old_price = old_price
                            updated = True
                        if variant.stock_quantity != stock_qty:
                            variant.stock_quantity = stock_qty
                            updated = True
                        if variant.is_in_stock != is_in_stock:
                            variant.is_in_stock = is_in_stock
                            updated = True
                        if variant.is_active != is_active:
                            variant.is_active = is_active
                            updated = True
                        if variant_image and variant.image != variant_image:
                            variant.image = variant_image
                            updated = True
                        if variant_title and variant.title != variant_title:
                            variant.title = variant_title
                            updated = True
                        
                        # Always update specifications, images, and subcategories (they might have changed)
                        # Check if there are any changes in these fields
                        has_spec_changes = False
                        has_image_changes = False
                        has_subcat_changes = False
                        
                        # We'll update these regardless, but mark if variant was updated
                        if updated:
                            variant.save()
                        
                        # Update other images (clear and recreate) - always do this
                        old_image_count = variant.images.count()
                        variant.images.all().delete()
                        new_image_count = 0
                        for i in range(1, 6):
                            img_col = f'other_image{i}'
                            if child_col_map.get(img_col) and row[child_col_map[img_col] - 1].value:
                                new_image_count += 1
                        
                        if old_image_count != new_image_count:
                            has_image_changes = True
                        
                        # Update subcategories - always do this
                        old_subcat_ids = set(variant.subcategories.values_list('id', flat=True))
                        variant.subcategories.clear()
                        new_subcat_ids = set()
                        for col_name, col_idx in child_col_map.items():
                            if col_name.startswith('Subcategory-'):
                                subcat_name = col_name.replace('Subcategory-', '').strip()
                                if row[col_idx - 1].value:
                                    val = str(row[col_idx - 1].value).strip().lower()
                                    if val == 'yes' or val == 'true' or val == '1':
                                        try:
                                            subcat = Subcategory.objects.get(name=subcat_name, category=product.category)
                                            variant.subcategories.add(subcat)
                                            new_subcat_ids.add(subcat.id)
                                        except Subcategory.DoesNotExist:
                                            pass
                        
                        if old_subcat_ids != new_subcat_ids:
                            has_subcat_changes = True
                        
                        # Update specifications (clear and recreate) - always do this
                        # Check if specs changed by comparing counts
                        old_spec_counts = {
                            'specifications': variant.specifications.count(),
                            'measurement_specs': variant.measurement_specs.count(),
                            'style_specs': variant.style_specs.count(),
                            'features': variant.features.count(),
                            'user_guide': variant.user_guide.count(),
                            'item_details': variant.item_details.count()
                        }
                        
                        spec_sections = {
                            'Specification:': ('specifications', ProductSpecification),
                            'Measurement Specification:': ('measurement_specs', VariantMeasurementSpec),
                            'Style Specification:': ('style_specs', VariantStyleSpec),
                            'Feature:': ('features', VariantFeature),
                            'User Guide:': ('user_guide', VariantUserGuide),
                            'Item Detail:': ('item_details', VariantItemDetail),
                        }
                        
                        for prefix, (section, model_class) in spec_sections.items():
                            # Clear existing specs for this section
                            getattr(variant, section).all().delete()
                            
                            # Add new specs from Excel
                            for col_name, col_idx in child_col_map.items():
                                if col_name.startswith(prefix):
                                    spec_name = col_name.replace(prefix, '').strip()
                                    spec_value = ''
                                    if row[col_idx - 1].value:
                                        spec_value = str(row[col_idx - 1].value).strip()
                                    
                                    if spec_value:
                                        model_class.objects.create(
                                            variant=variant,
                                            name=spec_name,
                                            value=spec_value,
                                            sort_order=0
                                        )
                        
                        # Check if spec counts changed
                        new_spec_counts = {
                            'specifications': variant.specifications.count(),
                            'measurement_specs': variant.measurement_specs.count(),
                            'style_specs': variant.style_specs.count(),
                            'features': variant.features.count(),
                            'user_guide': variant.user_guide.count(),
                            'item_details': variant.item_details.count()
                        }
                        
                        if old_spec_counts != new_spec_counts:
                            has_spec_changes = True
                        
                        # If any changes were made, mark as updated
                        if updated or has_spec_changes or has_image_changes or has_subcat_changes:
                            if not updated:  # Save if we haven't already
                                variant.save()
                            variants_updated += 1
                    else:
                        # Create new variant
                        variant = ProductVariant.objects.create(
                            product=product,
                            color=color,
                            size=size,
                            pattern=pattern,
                            quality=quality,
                            title=variant_title if variant_title else None,
                            price=price,
                            old_price=old_price,
                            stock_quantity=stock_qty,
                            is_in_stock=is_in_stock,
                            is_active=is_active,
                            image=variant_image if variant_image else None
                        )
                        variants_created += 1
                    
                    
                except Exception as e:
                    errors.append(f'Child row {row_num}: {str(e)}')
                    continue
            
            # Log action
            create_admin_log(
                request=request,
                action_type='update',
                model_name='Product',
                object_id=product.id,
                object_repr=str(product),
                details={
                    'action': 'update_from_excel',
                    'product_updated': product_updated,
                    'variants_created': variants_created,
                    'variants_updated': variants_updated
                }
            )
            
            # Build success message
            if product_updated or variants_updated > 0 or variants_created > 0:
                message = 'Product successfully edited'
            else:
                message = 'No changes detected'
            
            return Response({
                'success': True,
                'message': message,
                'product_updated': product_updated,
                'variants_created': variants_created,
                'variants_updated': variants_updated,
                'errors': errors[:50]
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            return Response(
                {'error': f'Failed to update product from Excel: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import products from Excel file with two-tab format (Parent Product and Child Variation)"""
        from openpyxl import load_workbook
        from io import BytesIO
        from django.utils.text import slugify
        from decimal import Decimal, InvalidOperation
        from products.models import (
            ProductSpecification, VariantMeasurementSpec, VariantStyleSpec,
            VariantFeature, VariantUserGuide, VariantItemDetail,
            ProductAboutItem, ProductVariantImage
        )
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Excel file is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        excel_file = request.FILES['file']
        
        try:
            # Load workbook
            wb = load_workbook(filename=BytesIO(excel_file.read()), data_only=True)
            
            # Check for required sheets
            if 'Parent Product' not in wb.sheetnames or 'Child Variation' not in wb.sheetnames:
                return Response(
                    {'error': 'Excel file must contain "Parent Product" and "Child Variation" sheets'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            ws_parent = wb['Parent Product']
            ws_child = wb['Child Variation']
            
            # Get headers from both sheets
            parent_headers = [cell.value for cell in ws_parent[1]]
            child_headers = [cell.value for cell in ws_child[1]]
            
            # Create column maps
            parent_col_map = {}
            for idx, header in enumerate(parent_headers, 1):
                if header:
                    parent_col_map[str(header).strip()] = idx
            
            child_col_map = {}
            for idx, header in enumerate(child_headers, 1):
                if header:
                    child_col_map[str(header).strip()] = idx
            
            # Validate required columns
            required_parent = ['Product Title*', 'SKU*', 'Short Description*', 'Category ID']
            required_child = ['Parent SKU*', 'Variant Color*', 'Variant Price*', 'Variant Stock Quantity*']
            
            missing_parent = [f for f in required_parent if f not in parent_col_map]
            missing_child = [f for f in required_child if f not in child_col_map]
            
            if missing_parent or missing_child:
                errors_list = []
                if missing_parent:
                    errors_list.append(f'Parent Product tab missing: {", ".join(missing_parent)}')
                if missing_child:
                    errors_list.append(f'Child Variation tab missing: {", ".join(missing_child)}')
                return Response(
                    {'error': '; '.join(errors_list)},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            products_created = 0
            variants_created = 0
            errors = []
            
            # Step 1: Read Parent Product tab and create products
            parent_products = {}  # {sku: {product_data, row_num}}
            
            for row_num, row in enumerate(ws_parent.iter_rows(min_row=2, values_only=False), start=2):
                # Skip empty rows
                if not row[parent_col_map['SKU*'] - 1].value:
                    continue
                
                try:
                    sku = str(row[parent_col_map['SKU*'] - 1].value).strip()
                    if not sku or sku == 'SKU*':
                        continue
                    
                    # Check if product with this SKU already exists
                    if Product.objects.filter(sku=sku).exists():
                        errors.append(f'Parent row {row_num}: Product with SKU "{sku}" already exists')
                        continue
                    
                    # Get category
                    category_id = None
                    if parent_col_map.get('Category ID'):
                        cat_id_val = row[parent_col_map['Category ID'] - 1].value
                        if cat_id_val:
                            try:
                                category_id = int(cat_id_val)
                            except (ValueError, TypeError):
                                errors.append(f'Parent row {row_num}: Invalid Category ID')
                                continue
                    
                    if not category_id:
                        errors.append(f'Parent row {row_num}: Category ID is required')
                        continue
                    
                    try:
                        category = Category.objects.get(id=category_id)
                    except Category.DoesNotExist:
                        errors.append(f'Parent row {row_num}: Category ID {category_id} not found')
                        continue
                    
                    # Extract product data
                    product_title = str(row[parent_col_map['Product Title*'] - 1].value).strip()
                    short_desc = str(row[parent_col_map['Short Description*'] - 1].value).strip()
                    long_desc = str(row[parent_col_map.get('Long Description', 0) - 1].value).strip() if parent_col_map.get('Long Description') and row[parent_col_map.get('Long Description', 0) - 1].value else ''
                    
                    # Material
                    material = None
                    if parent_col_map.get('Material') and row[parent_col_map['Material'] - 1].value:
                        material_name = str(row[parent_col_map['Material'] - 1].value).strip()
                        try:
                            material = Material.objects.get(name=material_name)
                        except Material.DoesNotExist:
                            pass
                    
                    # Other fields
                    brand = str(row[parent_col_map.get('Brand', 0) - 1].value).strip() if parent_col_map.get('Brand') and row[parent_col_map.get('Brand', 0) - 1].value else ''
                    dimensions = str(row[parent_col_map.get('Dimensions', 0) - 1].value).strip() if parent_col_map.get('Dimensions') and row[parent_col_map.get('Dimensions', 0) - 1].value else ''
                    weight = str(row[parent_col_map.get('Weight', 0) - 1].value).strip() if parent_col_map.get('Weight') and row[parent_col_map.get('Weight', 0) - 1].value else ''
                    warranty = str(row[parent_col_map.get('Warranty', 0) - 1].value).strip() if parent_col_map.get('Warranty') and row[parent_col_map.get('Warranty', 0) - 1].value else ''
                    
                    assembly_required = False
                    if parent_col_map.get('Assembly Required') and row[parent_col_map['Assembly Required'] - 1].value:
                        assembly_val = str(row[parent_col_map['Assembly Required'] - 1].value).strip().lower()
                        assembly_required = assembly_val == 'yes'
                    
                    estimated_delivery = 4
                    if parent_col_map.get('Estimated Delivery Days') and row[parent_col_map['Estimated Delivery Days'] - 1].value:
                        try:
                            estimated_delivery = int(row[parent_col_map['Estimated Delivery Days'] - 1].value)
                        except:
                            pass
                    
                    care_instructions = str(row[parent_col_map.get('Care Instructions', 0) - 1].value).strip() if parent_col_map.get('Care Instructions') and row[parent_col_map.get('Care Instructions', 0) - 1].value else ''
                    what_in_box = str(row[parent_col_map.get('What is in Box', 0) - 1].value).strip() if parent_col_map.get('What is in Box') and row[parent_col_map.get('What is in Box', 0) - 1].value else ''
                    
                    meta_title = str(row[parent_col_map.get('Meta Title', 0) - 1].value).strip() if parent_col_map.get('Meta Title') and row[parent_col_map.get('Meta Title', 0) - 1].value else ''
                    meta_desc = str(row[parent_col_map.get('Meta Description', 0) - 1].value).strip() if parent_col_map.get('Meta Description') and row[parent_col_map.get('Meta Description', 0) - 1].value else ''
                    
                    is_featured = False
                    if parent_col_map.get('Is Featured') and row[parent_col_map['Is Featured'] - 1].value:
                        featured_val = str(row[parent_col_map['Is Featured'] - 1].value).strip().lower()
                        is_featured = featured_val == 'yes'
                    
                    is_active = True
                    if parent_col_map.get('Is Active') and row[parent_col_map['Is Active'] - 1].value:
                        active_val = str(row[parent_col_map['Is Active'] - 1].value).strip().lower()
                        is_active = active_val == 'yes'
                    
                    # Generate slug
                    slug = slugify(product_title)
                    if Product.objects.filter(slug=slug).exists():
                        slug = f"{slug}-{timezone.now().strftime('%Y%m%d%H%M%S')}"
                    
                    # Create product
                    product = Product.objects.create(
                        title=product_title,
                        slug=slug,
                        sku=sku,
                        short_description=short_desc,
                        long_description=long_desc,
                        category=category,
                        material=material,
                        brand=brand,
                        dimensions=dimensions,
                        weight=weight,
                        warranty=warranty,
                        assembly_required=assembly_required,
                        estimated_delivery_days=estimated_delivery,
                        care_instructions=care_instructions,
                        what_in_box=what_in_box,
                        meta_title=meta_title,
                        meta_description=meta_desc,
                        is_featured=is_featured,
                        is_active=is_active
                    )
                    
                    # Process About Items (up to 10)
                    for i in range(1, 11):
                        col_name = f'About This Item {i}'
                        if parent_col_map.get(col_name) and row[parent_col_map[col_name] - 1].value:
                            item_text = str(row[parent_col_map[col_name] - 1].value).strip()
                            if item_text:
                                ProductAboutItem.objects.create(
                                    product=product,
                                    item=item_text,
                                    sort_order=i
                                )
                    
                    # Process Screen Offers (up to 10)
                    screen_offers = []
                    for i in range(1, 11):
                        title_col = f'Screen Offer Title {i}'
                        desc_col = f'Screen Offer Description {i}'
                        if parent_col_map.get(title_col) and row[parent_col_map[title_col] - 1].value:
                            offer_title = str(row[parent_col_map[title_col] - 1].value).strip()
                            offer_desc = ''
                            if parent_col_map.get(desc_col) and row[parent_col_map[desc_col] - 1].value:
                                offer_desc = str(row[parent_col_map[desc_col] - 1].value).strip()
                            if offer_title:
                                screen_offers.append({'title': offer_title, 'description': offer_desc})
                    
                    if screen_offers:
                        product.screen_offer = screen_offers
                        product.save()
                    
                    parent_products[sku] = {'product': product, 'category': category}
                    products_created += 1
                    
                except Exception as e:
                    errors.append(f'Parent row {row_num}: {str(e)}')
                    continue
            
            # Step 2: Read Child Variation tab and create variants
            for row_num, row in enumerate(ws_child.iter_rows(min_row=2, values_only=False), start=2):
                # Skip empty rows
                if not row[child_col_map['Parent SKU*'] - 1].value:
                    continue
                
                try:
                    parent_sku = str(row[child_col_map['Parent SKU*'] - 1].value).strip()
                    if not parent_sku or parent_sku == 'Parent SKU*':
                        continue
                    
                    # Find parent product
                    if parent_sku not in parent_products:
                        errors.append(f'Child row {row_num}: Parent SKU "{parent_sku}" not found in Parent Product tab')
                        continue
                    
                    product = parent_products[parent_sku]['product']
                    category = parent_products[parent_sku]['category']
                    
                    # Get color
                    color_name = str(row[child_col_map['Variant Color*'] - 1].value).strip()
                    try:
                        color = Color.objects.get(name=color_name)
                    except Color.DoesNotExist:
                        errors.append(f'Child row {row_num}: Color "{color_name}" not found')
                        continue
                    
                    # Variant fields
                    size = str(row[child_col_map.get('Variant Size', 0) - 1].value).strip() if child_col_map.get('Variant Size') and row[child_col_map.get('Variant Size', 0) - 1].value else ''
                    pattern = str(row[child_col_map.get('Variant Pattern', 0) - 1].value).strip() if child_col_map.get('Variant Pattern') and row[child_col_map.get('Variant Pattern', 0) - 1].value else ''
                    quality = str(row[child_col_map.get('Variant Quality', 0) - 1].value).strip() if child_col_map.get('Variant Quality') and row[child_col_map.get('Variant Quality', 0) - 1].value else ''
                    variant_title = str(row[child_col_map.get('Variant Title', 0) - 1].value).strip() if child_col_map.get('Variant Title') and row[child_col_map.get('Variant Title', 0) - 1].value else ''
                    
                    # Price
                    try:
                        price = Decimal(str(row[child_col_map['Variant Price*'] - 1].value))
                    except (ValueError, InvalidOperation, TypeError):
                        errors.append(f'Child row {row_num}: Invalid price value')
                        continue
                    
                    old_price = None
                    if child_col_map.get('Variant Old Price') and row[child_col_map['Variant Old Price'] - 1].value:
                        try:
                            old_price = Decimal(str(row[child_col_map['Variant Old Price'] - 1].value))
                        except:
                            pass
                    
                    # Stock
                    try:
                        stock_qty = int(row[child_col_map['Variant Stock Quantity*'] - 1].value or 0)
                    except (ValueError, TypeError):
                        errors.append(f'Child row {row_num}: Invalid stock quantity')
                        continue
                    
                    is_in_stock = True
                    if child_col_map.get('Variant Is In Stock') and row[child_col_map['Variant Is In Stock'] - 1].value:
                        stock_val = str(row[child_col_map['Variant Is In Stock'] - 1].value).strip().lower()
                        is_in_stock = stock_val == 'yes'
                    
                    is_active = True
                    if child_col_map.get('Variant Is Active') and row[child_col_map['Variant Is Active'] - 1].value:
                        active_val = str(row[child_col_map['Variant Is Active'] - 1].value).strip().lower()
                        is_active = active_val == 'yes'
                    
                    # Variant image
                    variant_image = ''
                    if child_col_map.get('Variant Image URL') and row[child_col_map['Variant Image URL'] - 1].value:
                        variant_image = str(row[child_col_map['Variant Image URL'] - 1].value).strip()
                    
                    # Create variant
                    variant = ProductVariant.objects.create(
                        product=product,
                        color=color,
                        size=size,
                        pattern=pattern,
                        quality=quality,
                        title=variant_title if variant_title else None,
                        price=price,
                        old_price=old_price,
                        stock_quantity=stock_qty,
                        is_in_stock=is_in_stock,
                        is_active=is_active,
                        image=variant_image if variant_image else None
                    )
                    
                    # Process other images (up to 5) - auto-generate alt_text and sort_order
                    image_counter = 0
                    for i in range(1, 6):
                        img_col = f'other_image{i}'
                        
                        if child_col_map.get(img_col) and row[child_col_map[img_col] - 1].value:
                            img_url = str(row[child_col_map[img_col] - 1].value).strip()
                            
                            if img_url:
                                image_counter += 1
                                # Auto-generate alt_text from image URL (extract filename or use default)
                                alt_text = ''
                                try:
                                    # Try to extract filename from URL
                                    parsed_url = urlparse(img_url)
                                    path = unquote(parsed_url.path)
                                    filename = path.split('/')[-1].split('.')[0] if '.' in path.split('/')[-1] else path.split('/')[-1]
                                    if filename:
                                        alt_text = filename.replace('_', ' ').replace('-', ' ').title()
                                except:
                                    pass
                                
                                # If alt_text is still empty, use default
                                if not alt_text:
                                    alt_text = f'Variant Image {image_counter}'
                                
                                # Auto-generate sort_order based on order of appearance
                                sort_order = image_counter
                                
                                ProductVariantImage.objects.create(
                                    variant=variant,
                                    image=img_url,
                                    alt_text=alt_text,
                                    sort_order=sort_order
                                )
                    
                    # Process subcategories (boolean columns like "Subcategory-{name}")
                    for col_name, col_idx in child_col_map.items():
                        if col_name.startswith('Subcategory-'):
                            subcat_name = col_name.replace('Subcategory-', '').strip()
                            if row[col_idx - 1].value:
                                val = str(row[col_idx - 1].value).strip().lower()
                                if val == 'yes' or val == 'true' or val == '1':
                                    try:
                                        subcat = Subcategory.objects.get(name=subcat_name, category=category)
                                        variant.subcategories.add(subcat)
                                    except Subcategory.DoesNotExist:
                                        pass
                    
                    # Process specifications
                    spec_sections = {
                        'Specification:': ('specifications', ProductSpecification),
                        'Measurement Specification:': ('measurement_specs', VariantMeasurementSpec),
                        'Style Specification:': ('style_specs', VariantStyleSpec),
                        'Feature:': ('features', VariantFeature),
                        'User Guide:': ('user_guide', VariantUserGuide),
                        'Item Detail:': ('item_details', VariantItemDetail),
                    }
                    
                    for col_name, col_idx in child_col_map.items():
                        for prefix, (section, model_class) in spec_sections.items():
                            if col_name.startswith(prefix):
                                spec_name = col_name.replace(prefix, '').strip()
                                spec_value = ''
                                if row[col_idx - 1].value:
                                    spec_value = str(row[col_idx - 1].value).strip()
                                
                                if spec_value:
                                    model_class.objects.create(
                                        variant=variant,
                                        name=spec_name,
                                        value=spec_value,
                                        sort_order=0
                                    )
                    
                    variants_created += 1
                    
                except Exception as e:
                    errors.append(f'Child row {row_num}: {str(e)}')
                    continue
            
            # Log action
            create_admin_log(
                request=request,
                action_type='bulk_create',
                model_name='Product',
                object_repr=f'Bulk import: {products_created} products, {variants_created} variants',
                details={
                    'products_created': products_created,
                    'variants_created': variants_created,
                    'errors_count': len(errors)
                }
            )
            
            return Response({
                'success': True,
                'message': f'Import completed: {products_created} products and {variants_created} variants created',
                'products_created': products_created,
                'variants_created': variants_created,
                'errors': errors[:50]  # Limit to first 50 errors
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            return Response(
                {'error': f'Failed to import Excel: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ==================== Order Management Views ====================
class AdminOrderViewSet(AdminLoggingMixin, viewsets.ReadOnlyModelViewSet):
    """Admin viewset for order management (read-only with custom actions)"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = Order.objects.all().select_related('user').prefetch_related(
        'items', 'status_history', 'notes'
    ).order_by('-created_at')
    serializer_class = AdminOrderListSerializer
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return AdminOrderDetailSerializer
        return AdminOrderListSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        status_filter = self.request.query_params.get('status', None)
        payment_status = self.request.query_params.get('payment_status', None)
        vendor = self.request.query_params.get('vendor', None)
        exclude_sixpine = self.request.query_params.get('exclude_sixpine', None)
        date_from = self.request.query_params.get('date_from', None)
        date_to = self.request.query_params.get('date_to', None)
        
        if search:
            queryset = queryset.filter(
                Q(user__username__icontains=search) |
                Q(user__email__icontains=search) |
                Q(user__mobile__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)
        if vendor:
            # If vendor is "sixpine", filter by brand="Sixpine" or vendor is None
            if vendor == 'sixpine':
                queryset = queryset.filter(
                    Q(items__product__brand__iexact='Sixpine') | Q(items__product__vendor__isnull=True)
                ).distinct()
            else:
                queryset = queryset.filter(items__vendor_id=vendor).distinct()
        elif exclude_sixpine == 'true':
            # Exclude Sixpine orders (orders with brand="Sixpine" or no vendor)
            queryset = queryset.exclude(
                Q(items__product__brand__iexact='Sixpine') | Q(items__product__vendor__isnull=True)
            ).distinct()
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """Update order status"""
        order = self.get_object()
        new_status = request.data.get('status')
        notes = request.data.get('notes', '')
        
        if not new_status:
            return Response(
                {'error': 'Status is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = order.status
        order.status = new_status
        order.save()
        
        # Create status history entry
        OrderStatusHistory.objects.create(
            order=order,
            status=new_status,
            notes=notes,
            created_by=request.user
        )
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='Order',
            object_id=order.id,
            object_repr=str(order),
            details={'status': {'old': old_status, 'new': new_status}}
        )
        
        serializer = self.get_serializer(order)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def update_payment_status(self, request, pk=None):
        """Update order payment status"""
        order = self.get_object()
        new_payment_status = request.data.get('payment_status')
        notes = request.data.get('notes', '')
        
        if not new_payment_status:
            return Response(
                {'error': 'Payment status is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_payment_status = order.payment_status
        order.payment_status = new_payment_status
        order.save()
        
        # Create order note if provided
        if notes:
            OrderNote.objects.create(
                order=order,
                content=notes,
                created_by=request.user
            )
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='Order',
            object_id=order.id,
            object_repr=str(order),
            details={'payment_status': {'old': old_payment_status, 'new': new_payment_status}}
        )
        
        serializer = self.get_serializer(order)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def update_tracking(self, request, pk=None):
        """Update order tracking information"""
        order = self.get_object()
        tracking_number = request.data.get('tracking_number')
        estimated_delivery = request.data.get('estimated_delivery')
        notes = request.data.get('notes', '')
        
        if tracking_number:
            order.tracking_number = tracking_number
        if estimated_delivery:
            order.estimated_delivery = estimated_delivery
        order.save()
        
        # Create order note if provided
        if notes:
            OrderNote.objects.create(
                order=order,
                content=notes,
                created_by=request.user
            )
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='Order',
            object_id=order.id,
            object_repr=str(order),
            details={'tracking_number': tracking_number, 'estimated_delivery': estimated_delivery}
        )
        
        serializer = self.get_serializer(order)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def notes(self, request, pk=None):
        """Get order notes"""
        order = self.get_object()
        notes = order.notes.all().order_by('-created_at')
        notes_data = [{
            'id': note.id,
            'note': note.note,
            'created_by': 'by Sixpine',  # Hardcoded as requested
            'created_at': note.created_at
        } for note in notes]
        return Response(notes_data)
    
    @action(detail=True, methods=['post'])
    def add_note(self, request, pk=None):
        """Add note to order"""
        order = self.get_object()
        note_text = request.data.get('note')
        
        if not note_text:
            return Response(
                {'error': 'Note text is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        note = OrderNote.objects.create(
            order=order,
            content=note_text,
            created_by=request.user
        )
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='Order',
            object_id=order.id,
            object_repr=str(order),
            details={'action': 'note_added', 'note_id': note.id}
        )
        
        return Response({
            'id': note.id,
            'note': note.note,
            'created_by': 'by Sixpine',  # Hardcoded as requested
            'created_at': note.created_at
        })


# ==================== Discount Management Views ====================
class AdminDiscountViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for discount management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = Discount.objects.all().order_by('-created_at')
    serializer_class = AdminDiscountSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset


# ==================== Coupon Management Views ====================
class AdminCouponViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for coupon management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = Coupon.objects.all().order_by('-created_at')
    serializer_class = AdminCouponSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset


# ==================== Payment Charges Settings ====================
@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated, IsAdminUser])
def payment_charges_settings(request):
    """Get or update payment charges settings"""
    def get_setting_value(key, default):
        value = GlobalSettings.get_setting(key, default)
        if isinstance(value, bool):
            return value
        if isinstance(value, str) and value.lower() in ['true', 'false']:
            return value.lower() == 'true'
        return value
    
    if request.method == 'GET':
        settings = {
            'platform_fee_upi': str(GlobalSettings.get_setting('platform_fee_upi', '0.00')),
            'platform_fee_card': str(GlobalSettings.get_setting('platform_fee_card', '2.36')),
            'platform_fee_netbanking': str(GlobalSettings.get_setting('platform_fee_netbanking', '2.36')),
            'platform_fee_cod': str(GlobalSettings.get_setting('platform_fee_cod', '0.00')),
            'tax_rate': str(GlobalSettings.get_setting('tax_rate', '5.00')),
            'razorpay_enabled': get_setting_value('razorpay_enabled', True),
            'cod_enabled': get_setting_value('cod_enabled', True),
        }
        serializer = PaymentChargeSerializer(settings)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        data = request.data
        
        # Update platform fees
        if 'platform_fee_upi' in data:
            GlobalSettings.set_setting('platform_fee_upi', data.get('platform_fee_upi', '0.00'), 'Platform fee percentage for UPI payments')
        if 'platform_fee_card' in data:
            GlobalSettings.set_setting('platform_fee_card', data.get('platform_fee_card', '2.36'), 'Platform fee percentage for Credit/Debit Card payments')
        if 'platform_fee_netbanking' in data:
            GlobalSettings.set_setting('platform_fee_netbanking', data.get('platform_fee_netbanking', '2.36'), 'Platform fee percentage for Net Banking payments')
        if 'platform_fee_cod' in data:
            GlobalSettings.set_setting('platform_fee_cod', data.get('platform_fee_cod', '0.00'), 'Platform fee percentage for COD payments')
        
        # Update tax rate
        if 'tax_rate' in data:
            GlobalSettings.set_setting('tax_rate', data.get('tax_rate', '5.00'), 'Tax rate percentage')
        
        # Update payment method enabled flags
        if 'razorpay_enabled' in data:
            GlobalSettings.set_setting('razorpay_enabled', data.get('razorpay_enabled', True), 'Enable Razorpay payment gateway')
        if 'cod_enabled' in data:
            GlobalSettings.set_setting('cod_enabled', data.get('cod_enabled', True), 'Enable Cash on Delivery')
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='GlobalSettings',
            object_id=None,
            object_repr='Payment Charges',
            details=data
        )
        
        # Return updated settings
        settings = {
            'platform_fee_upi': str(GlobalSettings.get_setting('platform_fee_upi', '0.00')),
            'platform_fee_card': str(GlobalSettings.get_setting('platform_fee_card', '2.36')),
            'platform_fee_netbanking': str(GlobalSettings.get_setting('platform_fee_netbanking', '2.36')),
            'platform_fee_cod': str(GlobalSettings.get_setting('platform_fee_cod', '0.00')),
            'tax_rate': str(GlobalSettings.get_setting('tax_rate', '5.00')),
            'razorpay_enabled': get_setting_value('razorpay_enabled', True),
            'cod_enabled': get_setting_value('cod_enabled', True),
        }
        serializer = PaymentChargeSerializer(settings)
        return Response(serializer.data)


# ==================== Global Settings ====================
@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated, IsAdminUser])
def global_settings(request):
    """Get or update global settings"""
    key = request.query_params.get('key', None)
    
    if request.method == 'GET':
        if key:
            try:
                setting = GlobalSettings.objects.get(key=key)
                serializer = GlobalSettingsSerializer(setting)
                return Response(serializer.data)
            except GlobalSettings.DoesNotExist:
                return Response(
                    {'error': 'Setting not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            settings = GlobalSettings.objects.all()
            serializer = GlobalSettingsSerializer(settings, many=True)
            return Response(serializer.data)
    
    elif request.method == 'PUT':
        data = request.data
        key = data.get('key')
        value = data.get('value')
        description = data.get('description', '')
        
        if not key:
            return Response(
                {'error': 'key is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Allow empty strings for optional fields (like social media URLs)
        if value is None:
            return Response(
                {'error': 'value is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        setting = GlobalSettings.set_setting(key, str(value), description)
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='GlobalSettings',
            object_id=setting.id,
            object_repr=str(setting),
            details={'key': key, 'value': value}
        )
        
        serializer = GlobalSettingsSerializer(setting)
        return Response(serializer.data)


# ==================== Contact Query Views ====================
class AdminContactQueryViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for contact query management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = ContactQuery.objects.all().order_by('-created_at')
    serializer_class = AdminContactQuerySerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        status_filter = self.request.query_params.get('status', None)
        
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone_number__icontains=search) |
                Q(message__icontains=search)
            )
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """Update contact query status"""
        contact_query = self.get_object()
        new_status = request.data.get('status')
        notes = request.data.get('notes', '')
        
        if not new_status:
            return Response(
                {'error': 'Status is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        contact_query.status = new_status
        if notes:
            contact_query.admin_notes = notes
        contact_query.save()
        
        serializer = self.get_serializer(contact_query)
        return Response(serializer.data)


# ==================== Bulk Order Views ====================
class AdminBulkOrderViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for bulk order management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = BulkOrder.objects.all().select_related('assigned_to').order_by('-created_at')
    serializer_class = AdminBulkOrderSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        status_filter = self.request.query_params.get('status', None)
        assigned_to = self.request.query_params.get('assigned_to', None)
        
        if search:
            queryset = queryset.filter(
                Q(company_name__icontains=search) |
                Q(contact_person__icontains=search) |
                Q(email__icontains=search) |
                Q(phone_number__icontains=search)
            )
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if assigned_to:
            queryset = queryset.filter(assigned_to_id=assigned_to)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """Update bulk order status"""
        bulk_order = self.get_object()
        status_value = request.data.get('status')
        notes = request.data.get('notes', '')
        quoted_price = request.data.get('quoted_price', None)
        
        if status_value not in dict(BulkOrder.STATUS_CHOICES).keys():
            return Response(
                {'error': 'Invalid status'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        bulk_order.status = status_value
        if notes:
            bulk_order.admin_notes = notes
        if quoted_price is not None:
            bulk_order.quoted_price = quoted_price
        bulk_order.save()
        
        serializer = self.get_serializer(bulk_order)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        """Assign bulk order to admin user"""
        bulk_order = self.get_object()
        admin_id = request.data.get('admin_id')
        
        try:
            admin_user = User.objects.get(id=admin_id, is_staff=True)
            bulk_order.assigned_to = admin_user
            bulk_order.save()
            
            serializer = self.get_serializer(bulk_order)
            return Response(serializer.data)
        except User.DoesNotExist:
            return Response(
                {'error': 'Admin user not found'},
                status=status.HTTP_404_NOT_FOUND
            )


# ==================== Admin Log Views ====================
class AdminLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Admin viewset for viewing logs (read-only)"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = AdminLog.objects.all().order_by('-created_at')
    serializer_class = AdminLogSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        action_type = self.request.query_params.get('action_type', None)
        model_name = self.request.query_params.get('model_name', None)
        user_id = self.request.query_params.get('user', None)
        search = self.request.query_params.get('search', None)
        date_from = self.request.query_params.get('date_from', None)
        date_to = self.request.query_params.get('date_to', None)
        
        if action_type:
            queryset = queryset.filter(action_type=action_type)
        if model_name:
            queryset = queryset.filter(model_name=model_name)
        if user_id:
            # Validate that user_id is a valid integer
            try:
                user_id_int = int(user_id)
                queryset = queryset.filter(user_id=user_id_int)
            except (ValueError, TypeError):
                # If not a valid integer, ignore the filter
                pass
        if search:
            queryset = queryset.filter(
                Q(user__email__icontains=search) |
                Q(user__username__icontains=search) |
                Q(model_name__icontains=search) |
                Q(object_repr__icontains=search) |
                Q(action_type__icontains=search)
            )
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """Override list to provide pagination"""
        from rest_framework.response import Response
        from rest_framework.pagination import PageNumberPagination
        
        class AdminLogPagination(PageNumberPagination):
            page_size = 20
            page_size_query_param = 'page_size'
            max_page_size = 100
        
        queryset = self.filter_queryset(self.get_queryset())
        paginator = AdminLogPagination()
        page = paginator.paginate_queryset(queryset, request)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# ==================== Home Page Content Views ====================
class AdminHomePageContentViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for managing home page content sections"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = HomePageContent.objects.all()
    serializer_class = HomePageContentSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        section_key = self.request.query_params.get('section_key', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if section_key:
            queryset = queryset.filter(section_key=section_key)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('order', 'section_name')
    
    def perform_create(self, serializer):
        """Create with logging"""
        instance = serializer.save()
        create_admin_log(
            request=self.request,
            action_type='create',
            model_name='HomePageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
    
    def perform_update(self, serializer):
        """Update with logging"""
        instance = serializer.save()
        create_admin_log(
            request=self.request,
            action_type='update',
            model_name='HomePageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
    
    def perform_destroy(self, instance):
        """Delete with logging"""
        create_admin_log(
            request=self.request,
            action_type='delete',
            model_name='HomePageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
        instance.delete()


# ==================== Bulk Order Page Content Views ====================
class AdminBulkOrderPageContentViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for managing bulk order page content sections"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = BulkOrderPageContent.objects.all()
    serializer_class = BulkOrderPageContentSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        section_key = self.request.query_params.get('section_key', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if section_key:
            queryset = queryset.filter(section_key=section_key)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('order', 'section_name')
    
    def perform_create(self, serializer):
        """Create with logging"""
        instance = serializer.save()
        create_admin_log(
            request=self.request,
            action_type='create',
            model_name='BulkOrderPageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
    
    def perform_update(self, serializer):
        """Update with logging"""
        instance = serializer.save()
        create_admin_log(
            request=self.request,
            action_type='update',
            model_name='BulkOrderPageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
    
    def perform_destroy(self, instance):
        """Delete with logging"""
        create_admin_log(
            request=self.request,
            action_type='delete',
            model_name='BulkOrderPageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
        instance.delete()


# ==================== FAQ Page Content Views ====================
class AdminFAQPageContentViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for managing FAQ page content sections"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = FAQPageContent.objects.all()
    serializer_class = FAQPageContentSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        section_key = self.request.query_params.get('section_key', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if section_key:
            queryset = queryset.filter(section_key=section_key)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('order', 'section_name')
    
    def perform_create(self, serializer):
        """Create with logging"""
        instance = serializer.save()
        create_admin_log(
            request=self.request,
            action_type='create',
            model_name='FAQPageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
    
    def perform_update(self, serializer):
        """Update with logging"""
        instance = serializer.save()
        create_admin_log(
            request=self.request,
            action_type='update',
            model_name='FAQPageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
    
    def perform_destroy(self, instance):
        """Delete with logging"""
        create_admin_log(
            request=self.request,
            action_type='delete',
            model_name='FAQPageContent',
            object_id=instance.id,
            object_repr=str(instance),
            details={'section_key': instance.section_key}
        )
        instance.delete()


# ==================== Advertisement Views ====================
class AdminAdvertisementViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for managing advertisements"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = Advertisement.objects.all()
    serializer_class = AdvertisementSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        is_active = self.request.query_params.get('is_active', None)
        
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('display_order', '-created_at')
    
    def perform_create(self, serializer):
        """Create with logging"""
        instance = serializer.save()
        create_admin_log(
            request=self.request,
            action_type='create',
            model_name='Advertisement',
            object_id=instance.id,
            object_repr=str(instance),
            details={'title': instance.title}
        )
    
    def perform_update(self, serializer):
        """Update with logging"""
        instance = serializer.save()
        create_admin_log(
            request=self.request,
            action_type='update',
            model_name='Advertisement',
            object_id=instance.id,
            object_repr=str(instance),
            details={'title': instance.title}
        )
    
    def perform_destroy(self, instance):
        """Delete with logging"""
        create_admin_log(
            request=self.request,
            action_type='delete',
            model_name='Advertisement',
            object_id=instance.id,
            object_repr=str(instance),
            details={'title': instance.title}
        )
        instance.delete()


class AdminDataRequestViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """ViewSet for managing data requests"""
    queryset = DataRequest.objects.all().order_by('-requested_at')
    serializer_class = AdminDataRequestSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter by status if provided
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        # Filter by request_type if provided
        request_type = self.request.query_params.get('request_type', None)
        if request_type:
            queryset = queryset.filter(request_type=request_type)
        return queryset
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a data request and generate Excel file"""
        data_request = self.get_object()
        
        if data_request.status != 'pending':
            return Response({
                'error': f'Request is already {data_request.status}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            import os
            from django.conf import settings
            
            # Generate file path
            file_name = f"{data_request.user.email}_{data_request.request_type}_{data_request.id}.xlsx"
            media_root = getattr(settings, 'MEDIA_ROOT', 'media')
            data_export_dir = os.path.join(media_root, 'data_exports')
            file_path = os.path.join(data_export_dir, file_name)
            
            # Generate Excel file based on request type
            if data_request.request_type == 'orders':
                export_orders_to_excel(data_request.user, file_path)
            elif data_request.request_type == 'addresses':
                export_addresses_to_excel(data_request.user, file_path)
            elif data_request.request_type == 'payment_options':
                export_payment_options_to_excel(data_request.user, file_path)
            
            # Update request status
            data_request.status = 'approved'
            data_request.approved_at = timezone.now()
            data_request.approved_by = request.user
            data_request.file_path = file_path
            data_request.save()
            
            # Log action
            create_admin_log(
                request=request,
                action_type='update',
                model_name='DataRequest',
                object_id=data_request.id,
                object_repr=str(data_request),
                details={'action': 'approve', 'request_type': data_request.request_type}
            )
            
            return Response({
                'success': True,
                'message': 'Request approved and file generated successfully',
                'data': AdminDataRequestSerializer(data_request).data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Failed to approve request: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a data request"""
        data_request = self.get_object()
        
        if data_request.status != 'pending':
            return Response({
                'error': f'Request is already {data_request.status}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        admin_notes = request.data.get('admin_notes', '')
        
        data_request.status = 'rejected'
        data_request.approved_by = request.user
        data_request.admin_notes = admin_notes
        data_request.save()
        
        # Log action
        create_admin_log(
            request=request,
            action_type='update',
            model_name='DataRequest',
            object_id=data_request.id,
            object_repr=str(data_request),
            details={'action': 'reject'}
        )
        
        return Response({
            'success': True,
            'message': 'Request rejected successfully',
            'data': AdminDataRequestSerializer(data_request).data
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download the generated Excel file"""
        data_request = self.get_object()
        
        if data_request.status != 'approved' and data_request.status != 'completed':
            return Response({
                'error': 'Request is not approved yet'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not data_request.file_path or not os.path.exists(data_request.file_path):
            return Response({
                'error': 'File not found. Please regenerate the file.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        from django.http import FileResponse
        
        file_name = os.path.basename(data_request.file_path)
        response = FileResponse(open(data_request.file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        
        # Mark as completed
        if data_request.status == 'approved':
            data_request.status = 'completed'
            data_request.completed_at = timezone.now()
            data_request.save()
        
        return response
    
    def perform_destroy(self, instance):
        """Override destroy to delete file and log action"""
        # Delete the Excel file if it exists
        if instance.file_path and os.path.exists(instance.file_path):
            try:
                os.remove(instance.file_path)
            except Exception as e:
                print(f"Error deleting file {instance.file_path}: {e}")
        
        # Log the deletion
        create_admin_log(
            request=self.request,
            action_type='delete',
            model_name='DataRequest',
            object_id=instance.id,
            object_repr=str(instance),
            details={'action': 'delete', 'request_type': instance.request_type, 'user_email': instance.user.email}
        )
        
        instance.delete()
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """Delete multiple data requests"""
        request_ids = request.data.get('ids', [])
        
        if not request_ids or not isinstance(request_ids, list):
            return Response({
                'error': 'Please provide a list of request IDs to delete'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        deleted_count = 0
        errors = []
        
        for request_id in request_ids:
            try:
                data_request = DataRequest.objects.get(id=request_id)
                
                # Delete the Excel file if it exists
                if data_request.file_path and os.path.exists(data_request.file_path):
                    try:
                        os.remove(data_request.file_path)
                    except Exception as e:
                        print(f"Error deleting file {data_request.file_path}: {e}")
                
                # Log the deletion
                create_admin_log(
                    request=request,
                    action_type='delete',
                    model_name='DataRequest',
                    object_id=data_request.id,
                    object_repr=str(data_request),
                    details={'action': 'bulk_delete', 'request_type': data_request.request_type, 'user_email': data_request.user.email}
                )
                
                data_request.delete()
                deleted_count += 1
            except DataRequest.DoesNotExist:
                errors.append(f"Request {request_id} not found")
            except Exception as e:
                errors.append(f"Error deleting request {request_id}: {str(e)}")
        
        response_data = {
            'success': True,
            'message': f'Successfully deleted {deleted_count} request(s)',
            'deleted_count': deleted_count
        }
        
        if errors:
            response_data['errors'] = errors
        
        return Response(response_data, status=status.HTTP_200_OK)


# ==================== Brand/Vendor Management Views ====================
class AdminBrandViewSet(AdminLoggingMixin, viewsets.ReadOnlyModelViewSet):
    """Admin viewset for brand/vendor management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    queryset = Vendor.objects.all().select_related('user').order_by('-created_at')
    serializer_class = AdminBrandSerializer
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return AdminBrandDetailSerializer
        return AdminBrandSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None)
        status_filter = self.request.query_params.get('status', None)
        is_verified = self.request.query_params.get('is_verified', None)
        
        if search:
            queryset = queryset.filter(
                Q(business_name__icontains=search) |
                Q(brand_name__icontains=search) |
                Q(business_email__icontains=search) |
                Q(user__email__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if is_verified is not None:
            queryset = queryset.filter(is_verified=is_verified.lower() == 'true')
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def suspend(self, request, pk=None):
        """Suspend a brand/vendor"""
        vendor = self.get_object()
        vendor.status = 'suspended'
        vendor.save()
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='Vendor',
            object_id=vendor.id,
            object_repr=str(vendor),
            details={'status': 'suspended'}
        )
        
        serializer = self.get_serializer(vendor)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Activate a brand/vendor"""
        vendor = self.get_object()
        vendor.status = 'active'
        vendor.is_verified = True
        vendor.save()
        
        create_admin_log(
            request=request,
            action_type='update',
            model_name='Vendor',
            object_id=vendor.id,
            object_repr=str(vendor),
            details={'status': 'active', 'is_verified': True}
        )
        
        serializer = self.get_serializer(vendor)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def products(self, request, pk=None):
        """Get all products for a specific brand"""
        vendor = self.get_object()
        from products.models import Product
        products = Product.objects.filter(vendor=vendor).select_related('category', 'subcategory')
        serializer = AdminProductListSerializer(products, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def orders(self, request, pk=None):
        """Get all orders for a specific brand"""
        vendor = self.get_object()
        from orders.models import Order, OrderItem
        orders = Order.objects.filter(items__vendor=vendor).distinct().order_by('-created_at')
        serializer = AdminOrderListSerializer(orders, many=True)
        return Response(serializer.data)


# ==================== Media Management Views ====================
class AdminMediaViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for media management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = AdminMediaSerializer
    
    def get_queryset(self):
        """Only show media uploaded by admin users (not sellers)"""
        queryset = Media.objects.filter(uploaded_by_user__isnull=False).order_by('-created_at')
        search = self.request.query_params.get('search', None)
        
        if search:
            queryset = queryset.filter(
                Q(file_name__icontains=search) |
                Q(alt_text__icontains=search) |
                Q(description__icontains=search) |
                Q(uploaded_by_user__email__icontains=search)
            )
        
        return queryset
    
    @action(detail=False, methods=['post'], url_path='upload')
    def upload_media(self, request):
        """Upload image to Cloudinary"""
        if 'image' not in request.FILES:
            return Response(
                {'error': 'No image file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        image_file = request.FILES['image']
        
        # Validate file type
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
        if image_file.content_type not in allowed_types:
            return Response(
                {'error': 'Invalid file type. Only images are allowed.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate file size (max 10MB)
        if image_file.size > 10 * 1024 * 1024:
            return Response(
                {'error': 'File size too large. Maximum size is 10MB.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            import cloudinary.uploader
            
            # Upload to Cloudinary
            upload_result = cloudinary.uploader.upload(
                image_file,
                folder='admin_media',
                resource_type='image'
            )
            
            # Create Media record
            media = Media(
                uploaded_by_user=request.user,
                cloudinary_url=upload_result['secure_url'],
                cloudinary_public_id=upload_result['public_id'],
                file_name=image_file.name,
                file_size=image_file.size,
                mime_type=image_file.content_type,
                alt_text=request.data.get('alt_text', ''),
                description=request.data.get('description', '')
            )
            media.full_clean()  # Validate model
            media.save()
            
            serializer = AdminMediaSerializer(media)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': f'Upload failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """Delete media from Cloudinary and database"""
        media = self.get_object()
        
        # Only allow deletion of own uploads
        if media.uploaded_by_user != request.user:
            return Response(
                {'error': 'You can only delete your own uploads'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            import cloudinary.uploader
            
            # Delete from Cloudinary if public_id exists
            if media.cloudinary_public_id:
                try:
                    cloudinary.uploader.destroy(media.cloudinary_public_id)
                except Exception as e:
                    # Log error but continue with database deletion
                    print(f"Cloudinary deletion error: {str(e)}")
            
            # Delete from database
            media.delete()
            
            return Response(status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            return Response(
                {'error': f'Deletion failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ==================== Packaging Feedback Views ====================
class AdminPackagingFeedbackViewSet(AdminLoggingMixin, viewsets.ModelViewSet):
    """Admin viewset for packaging feedback management"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = AdminPackagingFeedbackSerializer
    
    def get_queryset(self):
        queryset = PackagingFeedback.objects.all().select_related('user', 'reviewed_by').order_by('-created_at')
        search = self.request.query_params.get('search', None)
        status_filter = self.request.query_params.get('status', None)
        feedback_type = self.request.query_params.get('feedback_type', None)
        
        if search:
            queryset = queryset.filter(
                Q(message__icontains=search) |
                Q(user__email__icontains=search) |
                Q(email__icontains=search) |
                Q(name__icontains=search) |
                Q(order_id__icontains=search)
            )
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if feedback_type:
            queryset = queryset.filter(feedback_type=feedback_type)
        
        return queryset
    
    @action(detail=True, methods=['post'], url_path='update-status')
    def update_status(self, request, pk=None):
        """Update feedback status"""
        feedback = self.get_object()
        new_status = request.data.get('status')
        admin_notes = request.data.get('admin_notes', '')
        
        if new_status not in dict(PackagingFeedback.STATUS_CHOICES):
            return Response(
                {'error': 'Invalid status'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        feedback.status = new_status
        if admin_notes:
            feedback.admin_notes = admin_notes
        if new_status in ['reviewed', 'resolved']:
            feedback.reviewed_by = request.user
            feedback.reviewed_at = timezone.now()
        feedback.save()
        
        serializer = self.get_serializer(feedback)
        return Response(serializer.data)


# ==================== Product Review Management ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_review_list(request):
    """Get all product reviews (pending and approved) for admin"""
    from products.serializers import ProductReviewSerializer
    from admin_api.serializers import AdminProductReviewSerializer
    
    # Get filter parameters
    status_filter = request.query_params.get('status', 'all')  # 'all', 'pending', 'approved'
    product_id = request.query_params.get('product_id')
    vendor_id = request.query_params.get('vendor_id')
    
    queryset = ProductReview.objects.select_related('user', 'product', 'product__vendor').order_by('-created_at')
    
    # Apply filters
    if status_filter == 'pending':
        queryset = queryset.filter(is_approved=False)
    elif status_filter == 'approved':
        queryset = queryset.filter(is_approved=True)
    
    if product_id:
        queryset = queryset.filter(product_id=product_id)
    
    if vendor_id:
        queryset = queryset.filter(product__vendor_id=vendor_id)
    
    serializer = AdminProductReviewSerializer(queryset, many=True)
    return Response({
        'count': queryset.count(),
        'results': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_review_approve(request, review_id):
    """Approve a product review (admin can approve any review)"""
    try:
        review = ProductReview.objects.select_related('product', 'user').get(id=review_id)
        
        if review.is_approved:
            return Response({
                'success': False,
                'message': 'Review is already approved'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        review.is_approved = True
        review.save()
        
        # Update product rating statistics
        product = review.product
        approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
        if approved_reviews.exists():
            product.average_rating = approved_reviews.aggregate(
                avg_rating=Avg('rating')
            )['avg_rating'] or 0
            product.review_count = approved_reviews.count()
        else:
            product.average_rating = 0
            product.review_count = 0
        product.save()
        
        # Log admin action
        create_admin_log(
            request=request,
            action_type='approve_review',
            model_name='ProductReview',
            object_id=review.id,
            object_repr=f"Review for {product.title} by {review.user.email}",
            details={
                'product_id': product.id,
                'product_title': product.title,
                'user_email': review.user.email,
                'rating': review.rating
            }
        )
        
        return Response({
            'success': True,
            'message': 'Review approved successfully'
        })
        
    except ProductReview.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Review not found'
        }, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_review_reject(request, review_id):
    """Reject a product review (admin can reject any review)"""
    try:
        review = ProductReview.objects.select_related('product', 'user').get(id=review_id)
        
        if not review.is_approved:
            return Response({
                'success': False,
                'message': 'Review is already rejected'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        review.is_approved = False
        review.save()
        
        # Update product rating statistics
        product = review.product
        approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
        if approved_reviews.exists():
            product.average_rating = approved_reviews.aggregate(
                avg_rating=Avg('rating')
            )['avg_rating'] or 0
            product.review_count = approved_reviews.count()
        else:
            product.average_rating = 0
            product.review_count = 0
        product.save()
        
        # Log admin action
        create_admin_log(
            request=request,
            action_type='reject_review',
            model_name='ProductReview',
            object_id=review.id,
            object_repr=f"Review for {product.title} by {review.user.email}",
            details={
                'product_id': product.id,
                'product_title': product.title,
                'user_email': review.user.email,
                'rating': review.rating
            }
        )
        
        return Response({
            'success': True,
            'message': 'Review rejected successfully'
        })
        
    except ProductReview.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Review not found'
        }, status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_review_delete(request, review_id):
    """Delete a product review (admin can delete any review)"""
    try:
        review = ProductReview.objects.select_related('product', 'user').get(id=review_id)
        product = review.product
        user_email = review.user.email
        rating = review.rating
        
        # Delete the review
        review.delete()
        
        # Update product rating statistics
        approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
        if approved_reviews.exists():
            product.average_rating = approved_reviews.aggregate(
                avg_rating=Avg('rating')
            )['avg_rating'] or 0
            product.review_count = approved_reviews.count()
        else:
            product.average_rating = 0
            product.review_count = 0
        product.save()
        
        # Log admin action
        create_admin_log(
            request=request,
            action_type='delete_review',
            model_name='ProductReview',
            object_id=review_id,
            object_repr=f"Review for {product.title} by {user_email}",
            details={
                'product_id': product.id,
                'product_title': product.title,
                'user_email': user_email,
                'rating': rating
            }
        )
        
        return Response({
            'success': True,
            'message': 'Review deleted successfully'
        })
        
    except ProductReview.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Review not found'
        }, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_review_delete_all(request):
    """Delete all reviews (admin only)"""
    try:
        # Get filter parameters
        status_filter = request.data.get('status', 'all')
        
        # Build query
        queryset = ProductReview.objects.all()
        
        if status_filter == 'pending':
            queryset = queryset.filter(is_approved=False)
        elif status_filter == 'approved':
            queryset = queryset.filter(is_approved=True)
        
        # Get count before deletion
        count = queryset.count()
        
        # Get product IDs that will be affected
        affected_products = set(queryset.values_list('product_id', flat=True))
        
        # Delete reviews
        queryset.delete()
        
        # Update product rating statistics for affected products
        from products.models import Product
        for product_id in affected_products:
            product = Product.objects.get(id=product_id)
            approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
            if approved_reviews.exists():
                product.average_rating = approved_reviews.aggregate(
                    avg_rating=Avg('rating')
                )['avg_rating'] or 0
                product.review_count = approved_reviews.count()
            else:
                product.average_rating = 0
                product.review_count = 0
            product.save()
        
        # Log admin action
        create_admin_log(
            request=request,
            action_type='delete_all_reviews',
            model_name='ProductReview',
            object_repr=f"Deleted {count} reviews",
            details={
                'count': count,
                'status_filter': status_filter
            }
        )
        
        return Response({
            'success': True,
            'message': f'{count} review(s) deleted successfully'
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Failed to delete reviews: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
