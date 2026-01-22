"""Utility functions for Excel-based bulk product creation"""
import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from products.models import Category, Subcategory, Color, Material, CategorySpecificationTemplate, Product, ProductVariant


def generate_product_template(category_id):
    """Generate category-specific Excel template for bulk product creation"""
    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        raise ValueError(f"Category with id {category_id} does not exist")
    
    # Get subcategories for this category
    subcategories = Subcategory.objects.filter(category=category, is_active=True).order_by('name')
    subcategory_names = [sub.name for sub in subcategories]
    subcategory_list = list(subcategories)  # Keep list for column generation
    
    # Get category specification templates
    spec_templates = CategorySpecificationTemplate.objects.filter(
        category=category,
        is_active=True
    ).order_by('section', 'sort_order', 'field_name')
    
    # Group specifications by section
    spec_groups = {
        'specifications': [],
        'measurement_specs': [],
        'style_specs': [],
        'features': [],
        'user_guide': [],
        'item_details': []
    }
    
    for template in spec_templates:
        if template.section in spec_groups:
            spec_groups[template.section].append(template.field_name)
    
    # If no category templates exist, use default specification names
    if not spec_templates.exists():
        # Default specification names for each section
        spec_groups['specifications'] = ['Brand', 'Model', 'Material', 'Color', 'Weight']
        spec_groups['measurement_specs'] = ['Length', 'Width', 'Height', 'Weight', 'Dimensions']
        spec_groups['style_specs'] = ['Style', 'Pattern', 'Finish', 'Design']
        spec_groups['features'] = ['Feature 1', 'Feature 2', 'Feature 3']
        spec_groups['user_guide'] = ['Assembly', 'Care Instructions', 'Usage']
        spec_groups['item_details'] = ['Warranty', 'Package Contents', 'Additional Info']
    
    # Get all colors and materials for dropdowns
    colors = Color.objects.filter(is_active=True).order_by('name')
    color_names = [color.name for color in colors]
    
    materials = Material.objects.filter(is_active=True).order_by('name')
    material_names = [material.name for material in materials]
    
    # Create workbook with two tabs
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ========== PARENT PRODUCT TAB ==========
    ws_parent = wb.create_sheet("Parent Product")
    
    # Parent headers
    parent_headers = [
        'Variation Relationship',
        'Product Title*',
        'SKU*',
        'Price',
        'Old Price',
        'Short Description*',
        'Long Description',
        'Category*',
        'Category ID',
        'Material',
        'Brand',
        'Dimensions',
        'Weight',
        'Warranty',
        'Assembly Required',
        'Estimated Delivery Days',
        'Meta Title',
        'Meta Description',
        'Is Featured',
        'Is Active',
        'Care Instructions',
        'What is in Box',
        'Parent Main Image URL',
    ]
    
    # Add About Items columns (up to 10) - just description, no title
    for i in range(1, 11):
        parent_headers.append(f'About This Item {i}')
    
    # Add Screen Offer columns (up to 10) - has title and description
    for i in range(1, 11):
        parent_headers.append(f'Screen Offer Title {i}')
        parent_headers.append(f'Screen Offer Description {i}')
    
    # Define recommendation types (used in multiple places)
    recommendation_types = [
        ('buy_with', 'Buy it with'),
        ('frequently_viewed', 'Frequently viewed'),
        ('inspired_by', 'Inspired by browsing history'),
        ('similar', 'Similar products'),
        ('recommended', 'Recommended for you')
    ]
    
    # Add Product Recommendation columns (5 columns per type, can be extended manually)
    # Each column contains SKU dropdown (can also manually enter SKU)
    for rec_type, rec_label in recommendation_types:
        for i in range(1, 6):  # 5 columns per type
            parent_headers.append(f'{rec_label} Product {i} SKU')
    
    # Note: Specification columns are NOT in Parent tab - they are variant-specific and only in Child Variation tab
    
    # Style parent header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(parent_headers, 1):
        cell = ws_parent.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws_parent.freeze_panes = 'A2'
    
    # Add sample row
    row_num = 2
    col_idx = 1
    ws_parent.cell(row=row_num, column=col_idx, value='Parent'); col_idx += 1  # Variation Relationship
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Product Title
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # SKU
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Price
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Old Price
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Short Description
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Long Description
    ws_parent.cell(row=row_num, column=col_idx, value=category.name); col_idx += 1  # Category
    ws_parent.cell(row=row_num, column=col_idx, value=category.id); col_idx += 1  # Category ID
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Material (removed Subcategory* column)
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Brand
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Dimensions
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Weight
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Warranty
    ws_parent.cell(row=row_num, column=col_idx, value='No'); col_idx += 1  # Assembly Required
    ws_parent.cell(row=row_num, column=col_idx, value=4); col_idx += 1  # Estimated Delivery Days
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Meta Title
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Meta Description
    ws_parent.cell(row=row_num, column=col_idx, value='No'); col_idx += 1  # Is Featured
    ws_parent.cell(row=row_num, column=col_idx, value='Yes'); col_idx += 1  # Is Active
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Care Instructions
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # What is in Box
    ws_parent.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Parent Main Image URL
    
    # About Items columns (all empty for template)
    for i in range(10):
        ws_parent.cell(row=row_num, column=col_idx, value='')
        col_idx += 1
    
    # Screen Offer columns (all empty for template)
    for i in range(10):
        ws_parent.cell(row=row_num, column=col_idx, value='')  # Title
        col_idx += 1
        ws_parent.cell(row=row_num, column=col_idx, value='')  # Description
        col_idx += 1
    
    # Product Recommendation columns (all empty for template)
    # Each column contains SKU (dropdown or manual entry)
    for rec_type, rec_label in recommendation_types:
        for i in range(5):  # 5 columns per type
            ws_parent.cell(row=row_num, column=col_idx, value='')  # Product SKU
            col_idx += 1
    
    # Note: Specifications are variant-specific and are NOT in Parent tab - they are only in Child Variation tab
    # Note: Recommendation columns do not have dropdown - users manually enter SKUs
    
    # ========== CHILD VARIATION TAB ==========
    ws_child = wb.create_sheet("Child Variation")
    
    # Child headers
    child_headers = [
        'Variation Relationship',
        'Parent SKU*',
        'Variant Color*',
        'Variant Size',
        'Variant Pattern',
        'Variant Quality',
        'Variant Title',
        'Variant SKU',
        'Variant Price*',
        'Variant Old Price',
        'Variant Stock Quantity*',
        'Variant Is In Stock',
        'Variant Is Active',
        'Variant Image URL',
        'Variant Video URL',
    ]
    
    # Add other_image columns (up to 9) - alt text and sort order are auto-generated during import
    for i in range(1, 10):
        child_headers.append(f'other_image{i}')
    
    # Add subcategory boolean columns for variants (one per subcategory)
    for subcategory in subcategory_list:
        child_headers.append(f'Subcategory-{subcategory.name}')
    
    # Add category template specification columns to child (from category templates)
    # Format: "Specification: {name}", "Style Specification: {name}", etc.
    # User only enters values, not names
    for spec_name in spec_groups['specifications']:
        child_headers.append(f'Specification: {spec_name}')
    for spec_name in spec_groups['measurement_specs']:
        child_headers.append(f'Measurement Specification: {spec_name}')
    for spec_name in spec_groups['style_specs']:
        child_headers.append(f'Style Specification: {spec_name}')
    for spec_name in spec_groups['features']:
        child_headers.append(f'Feature: {spec_name}')
    for spec_name in spec_groups['user_guide']:
        child_headers.append(f'User Guide: {spec_name}')
    for spec_name in spec_groups['item_details']:
        child_headers.append(f'Item Detail: {spec_name}')
    
    # Style child header
    for col_num, header in enumerate(child_headers, 1):
        cell = ws_child.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws_child.freeze_panes = 'A2'
    
    # Add sample row
    row_num = 2
    col_idx = 1
    ws_child.cell(row=row_num, column=col_idx, value='Child'); col_idx += 1  # Variation Relationship
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Parent SKU
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Color
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Size
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Pattern
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Quality
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Title
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant SKU
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Price
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Old Price
    ws_child.cell(row=row_num, column=col_idx, value=0); col_idx += 1  # Variant Stock Quantity
    ws_child.cell(row=row_num, column=col_idx, value='Yes'); col_idx += 1  # Variant Is In Stock
    ws_child.cell(row=row_num, column=col_idx, value='Yes'); col_idx += 1  # Variant Is Active
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Image URL
    ws_child.cell(row=row_num, column=col_idx, value=''); col_idx += 1  # Variant Video URL
    
    # Other images (all empty for template) - alt text and sort order are auto-generated during import (max 9)
    for i in range(9):
        ws_child.cell(row=row_num, column=col_idx, value='')  # Image URL
        col_idx += 1
    
    # Subcategory boolean columns (all No for template)
    for subcategory in subcategory_list:
        ws_child.cell(row=row_num, column=col_idx, value='No')
        col_idx += 1
    
    # Leave specification columns empty for template
    for col_idx in range(col_idx, len(child_headers) + 1):
        ws_child.cell(row=row_num, column=col_idx, value='')
    
    # Add data validations
    # Parent sheet validations
    material_col = get_column_letter(8)
    assembly_col = get_column_letter(13)
    featured_col = get_column_letter(17)
    active_col = get_column_letter(18)
    subcategory_start_col = 22  # After "What is in Box"
    
    if material_names:
        material_dv = DataValidation(type="list", formula1=f'"{",".join(material_names)}"', allow_blank=True)
        ws_parent.add_data_validation(material_dv)
        material_dv.add(f'{material_col}2:{material_col}1000')
    
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws_parent.add_data_validation(yes_no_dv)
    yes_no_dv.add(f'{assembly_col}2:{assembly_col}1000')
    yes_no_dv.add(f'{featured_col}2:{featured_col}1000')
    yes_no_dv.add(f'{active_col}2:{active_col}1000')
    
    # Add Yes/No validation for subcategory boolean columns
    for idx, subcategory in enumerate(subcategory_list):
        subcategory_col = get_column_letter(subcategory_start_col + idx)
        yes_no_dv.add(f'{subcategory_col}2:{subcategory_col}1000')
    
    # Note: Recommendation columns do not have dropdown validation
    # Users can manually enter SKU values in the recommendation columns
    
    # Child sheet validations
    color_col = get_column_letter(3)
    stock_status_col = get_column_letter(11)
    variant_subcategory_start_col = 37  # After Variant Image URL (13) + 9 other_image columns (9 columns: 9 images * 1 column each)
    
    if color_names:
        color_dv = DataValidation(type="list", formula1=f'"{",".join(color_names)}"', allow_blank=False)
        ws_child.add_data_validation(color_dv)
        color_dv.add(f'{color_col}2:{color_col}1000')
    
    yes_no_dv_child = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws_child.add_data_validation(yes_no_dv_child)
    yes_no_dv_child.add(f'{stock_status_col}2:{stock_status_col}1000')
    
    # Add Yes/No validation for variant subcategory boolean columns
    for idx, subcategory in enumerate(subcategory_list):
        subcategory_col = get_column_letter(variant_subcategory_start_col + idx)
        yes_no_dv_child.add(f'{subcategory_col}2:{subcategory_col}1000')
    
    # Auto-adjust column widths for both sheets
    for ws in [ws_parent, ws_child]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 15), 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    return wb


def export_product_to_excel(product_id):
    """Export existing product with all variants to Excel in two tabs: Parent and Child"""
    try:
        product = Product.objects.prefetch_related(
            'variants__color',
            'variants__subcategories',
            'variants__specifications',
            'variants__measurement_specs',
            'variants__style_specs',
            'variants__features',
            'variants__user_guide',
            'variants__item_details',
            'variants__images',
            'about_items',
            'subcategories'
        ).get(id=product_id)
    except Product.DoesNotExist:
        raise ValueError(f"Product with id {product_id} does not exist")
    
    category = product.category
    
    # Get subcategories for this category
    subcategories = Subcategory.objects.filter(category=category, is_active=True).order_by('name')
    subcategory_names = [sub.name for sub in subcategories]
    subcategory_list = list(subcategories)  # Keep list for column generation
    
    # Get category specification templates
    spec_templates = CategorySpecificationTemplate.objects.filter(
        category=category,
        is_active=True
    ).order_by('section', 'sort_order', 'field_name')
    
    # Group specifications by section
    spec_groups = {
        'specifications': [],
        'measurement_specs': [],
        'style_specs': [],
        'features': [],
        'user_guide': [],
        'item_details': []
    }
    
    for template in spec_templates:
        if template.section in spec_groups:
            spec_groups[template.section].append(template.field_name)
    
    # Get all colors and materials for dropdowns
    colors = Color.objects.filter(is_active=True).order_by('name')
    color_names = [color.name for color in colors]
    
    materials = Material.objects.filter(is_active=True).order_by('name')
    material_names = [material.name for material in materials]
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ========== PARENT PRODUCT TAB ==========
    ws_parent = wb.create_sheet("Parent Product")
    
    # Parent headers
    parent_headers = [
        'Variation Relationship',
        'Product Title*',
        'SKU*',
        'Price',
        'Old Price',
        'Short Description*',
        'Long Description',
        'Category*',
        'Category ID',
        'Material',
        'Brand',
        'Dimensions',
        'Weight',
        'Warranty',
        'Assembly Required',
        'Estimated Delivery Days',
        'Meta Title',
        'Meta Description',
        'Is Featured',
        'Is Active',
        'Care Instructions',
        'What is in Box',
        'Parent Main Image URL',
    ]
    
    # Add About Items columns (up to 10) - just description, no title
    for i in range(1, 11):
        parent_headers.append(f'About This Item {i}')
    
    # Add Screen Offer columns (up to 10) - has title and description
    for i in range(1, 11):
        parent_headers.append(f'Screen Offer Title {i}')
        parent_headers.append(f'Screen Offer Description {i}')
    
    # Define recommendation types (used in multiple places)
    recommendation_types = [
        ('buy_with', 'Buy it with'),
        ('frequently_viewed', 'Frequently viewed'),
        ('inspired_by', 'Inspired by browsing history'),
        ('similar', 'Similar products'),
        ('recommended', 'Recommended for you')
    ]
    
    # Add Product Recommendation columns (5 columns per type, can be extended manually)
    # Each column contains SKU dropdown (can also manually enter SKU)
    for rec_type, rec_label in recommendation_types:
        for i in range(1, 6):  # 5 columns per type
            parent_headers.append(f'{rec_label} Product {i} SKU')
    
    # Note: Specification columns are NOT in Parent tab - they are variant-specific and only in Child Variation tab
    
    # Style parent header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(parent_headers, 1):
        cell = ws_parent.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws_parent.freeze_panes = 'A2'
    
    # Write parent product data
    row_num = 2
    col_idx = 1
    ws_parent.cell(row=row_num, column=col_idx, value='Parent'); col_idx += 1  # Variation Relationship
    ws_parent.cell(row=row_num, column=col_idx, value=product.title); col_idx += 1  # Product Title
    ws_parent.cell(row=row_num, column=col_idx, value=product.sku or ''); col_idx += 1  # SKU
    # Use first active variant for price fields
    first_variant = product.variants.filter(is_active=True).first()
    ws_parent.cell(row=row_num, column=col_idx, value=float(first_variant.price) if first_variant and first_variant.price is not None else ''); col_idx += 1  # Price
    ws_parent.cell(row=row_num, column=col_idx, value=float(first_variant.old_price) if first_variant and first_variant.old_price is not None else ''); col_idx += 1  # Old Price
    ws_parent.cell(row=row_num, column=col_idx, value=product.short_description); col_idx += 1  # Short Description
    ws_parent.cell(row=row_num, column=col_idx, value=product.long_description or ''); col_idx += 1  # Long Description
    ws_parent.cell(row=row_num, column=col_idx, value=category.name); col_idx += 1  # Category
    ws_parent.cell(row=row_num, column=col_idx, value=category.id); col_idx += 1  # Category ID
    ws_parent.cell(row=row_num, column=col_idx, value=product.material.name if product.material else ''); col_idx += 1  # Material (removed Subcategory* column)
    ws_parent.cell(row=row_num, column=col_idx, value=product.brand or ''); col_idx += 1  # Brand
    ws_parent.cell(row=row_num, column=col_idx, value=product.dimensions or ''); col_idx += 1  # Dimensions
    ws_parent.cell(row=row_num, column=col_idx, value=product.weight or ''); col_idx += 1  # Weight
    ws_parent.cell(row=row_num, column=col_idx, value=product.warranty or ''); col_idx += 1  # Warranty
    ws_parent.cell(row=row_num, column=col_idx, value='Yes' if product.assembly_required else 'No'); col_idx += 1  # Assembly Required
    ws_parent.cell(row=row_num, column=col_idx, value=product.estimated_delivery_days); col_idx += 1  # Estimated Delivery Days
    ws_parent.cell(row=row_num, column=col_idx, value=product.meta_title or ''); col_idx += 1  # Meta Title
    ws_parent.cell(row=row_num, column=col_idx, value=product.meta_description or ''); col_idx += 1  # Meta Description
    ws_parent.cell(row=row_num, column=col_idx, value='Yes' if product.is_featured else 'No'); col_idx += 1  # Is Featured
    ws_parent.cell(row=row_num, column=col_idx, value='Yes' if product.is_active else 'No'); col_idx += 1  # Is Active
    ws_parent.cell(row=row_num, column=col_idx, value=product.care_instructions or ''); col_idx += 1  # Care Instructions
    ws_parent.cell(row=row_num, column=col_idx, value=product.what_in_box or ''); col_idx += 1  # What is in Box
    ws_parent.cell(row=row_num, column=col_idx, value=product.parent_main_image or ''); col_idx += 1  # Parent Main Image URL
    
    # About Items (up to 10)
    about_items = product.about_items.all().order_by('sort_order')[:10]
    about_items_list = [item.item for item in about_items]
    for idx in range(10):
        ws_parent.cell(row=row_num, column=col_idx, value=about_items_list[idx] if idx < len(about_items_list) else '')
        col_idx += 1
    
    # Screen Offers (up to 10)
    screen_offers = product.screen_offer if isinstance(product.screen_offer, list) else []
    screen_offers_list = []
    for offer in screen_offers[:10]:
        if isinstance(offer, dict):
            screen_offers_list.append({'title': offer.get('title', ''), 'description': offer.get('description', '')})
        else:
            screen_offers_list.append({'title': str(offer), 'description': ''})
    
    for idx in range(10):
        if idx < len(screen_offers_list):
            ws_parent.cell(row=row_num, column=col_idx, value=screen_offers_list[idx].get('title', ''))
            col_idx += 1
            ws_parent.cell(row=row_num, column=col_idx, value=screen_offers_list[idx].get('description', ''))
            col_idx += 1
        else:
             ws_parent.cell(row=row_num, column=col_idx, value='')
             col_idx += 1
             ws_parent.cell(row=row_num, column=col_idx, value='')
             col_idx += 1
     
    # Product Recommendations (up to 5 per type, can be extended manually)
    recommendation_types = [
        ('buy_with', 'Buy it with'),
        ('frequently_viewed', 'Frequently viewed'),
        ('inspired_by', 'Inspired by browsing history'),
        ('similar', 'Similar products'),
        ('recommended', 'Recommended for you')
    ]
    
    for rec_type, rec_label in recommendation_types:
        recommendations = product.recommendations.filter(
            recommendation_type=rec_type,
            is_active=True
        ).select_related('recommended_product').order_by('sort_order')[:5]
        
        for idx in range(5):
            if idx < len(recommendations):
                rec = recommendations[idx]
                recommended_product = rec.recommended_product
                ws_parent.cell(row=row_num, column=col_idx, value=recommended_product.sku or '')
                col_idx += 1
            else:
                ws_parent.cell(row=row_num, column=col_idx, value='')
                col_idx += 1
     
     # Note: Specifications are variant-specific and are NOT in Parent tab - they are only in Child Variation tab
     
     # ========== CHILD VARIATION TAB ==========
    ws_child = wb.create_sheet("Child Variation")
    
    # Child headers
    child_headers = [
        'Variation Relationship',
        'Parent SKU*',
        'Variant Color*',
        'Variant Size',
        'Variant Pattern',
        'Variant Quality',
        'Variant Title',
        'Variant SKU',
        'Variant Price*',
        'Variant Old Price',
        'Variant Stock Quantity*',
        'Variant Is In Stock',
        'Variant Is Active',
        'Variant Image URL',
        'Variant Video URL',
    ]
    
    # Add other_image columns (up to 9) - alt text and sort order are auto-generated during import
    for i in range(1, 10):
        child_headers.append(f'other_image{i}')
    
    # Add subcategory boolean columns for variants (one per subcategory)
    for subcategory in subcategory_list:
        child_headers.append(f'Subcategory-{subcategory.name}')
    
    # Add category template specification columns to child (from category templates or defaults)
    # Format: "Specification: {name}", "Style Specification: {name}", etc.
    # User only enters values, not names
    for spec_name in spec_groups['specifications']:
        child_headers.append(f'Specification: {spec_name}')
    for spec_name in spec_groups['measurement_specs']:
        child_headers.append(f'Measurement Specification: {spec_name}')
    for spec_name in spec_groups['style_specs']:
        child_headers.append(f'Style Specification: {spec_name}')
    for spec_name in spec_groups['features']:
        child_headers.append(f'Feature: {spec_name}')
    for spec_name in spec_groups['user_guide']:
        child_headers.append(f'User Guide: {spec_name}')
    for spec_name in spec_groups['item_details']:
        child_headers.append(f'Item Detail: {spec_name}')
    
    # Collect all unique specification names from all variants (for columns not in templates)
    all_variant_spec_names = {
        'specifications': set(),
        'measurement_specs': set(),
        'style_specs': set(),
        'features': set(),
        'user_guide': set(),
        'item_details': set()
    }
    
    # For item_details and user_guide, count max occurrences per name across all variants
    item_details_name_max_count = {}  # name -> max count across all variants
    user_guide_name_max_count = {}  # name -> max count across all variants
    
    variants = product.variants.all()
    for variant in variants:
        # Get all specification names from this variant
        for spec in variant.specifications.all():
            all_variant_spec_names['specifications'].add(spec.name)
        for spec in variant.measurement_specs.all():
            all_variant_spec_names['measurement_specs'].add(spec.name)
        for spec in variant.style_specs.all():
            all_variant_spec_names['style_specs'].add(spec.name)
        for spec in variant.features.all():
            all_variant_spec_names['features'].add(spec.name)
        
        # For user_guide and item_details, count occurrences per name per variant
        user_guide_counts = {}
        for spec in variant.user_guide.all().order_by('sort_order', 'id'):
            all_variant_spec_names['user_guide'].add(spec.name)
            user_guide_counts[spec.name] = user_guide_counts.get(spec.name, 0) + 1
            # Update max count
            if spec.name not in user_guide_name_max_count or user_guide_counts[spec.name] > user_guide_name_max_count[spec.name]:
                user_guide_name_max_count[spec.name] = user_guide_counts[spec.name]
        
        item_details_counts = {}
        for spec in variant.item_details.all().order_by('sort_order', 'id'):
            all_variant_spec_names['item_details'].add(spec.name)
            item_details_counts[spec.name] = item_details_counts.get(spec.name, 0) + 1
            # Update max count
            if spec.name not in item_details_name_max_count or item_details_counts[spec.name] > item_details_name_max_count[spec.name]:
                item_details_name_max_count[spec.name] = item_details_counts[spec.name]
    
    # Add variant-specific specifications that are not in category templates
    template_spec_names = {
        'specifications': set(spec_groups['specifications']),
        'measurement_specs': set(spec_groups['measurement_specs']),
        'style_specs': set(spec_groups['style_specs']),
        'features': set(spec_groups['features']),
        'user_guide': set(spec_groups['user_guide']),
        'item_details': set(spec_groups['item_details'])
    }
    
    # Add additional columns for variant specs not in templates
    additional_specs = {
        'specifications': sorted(all_variant_spec_names['specifications'] - template_spec_names['specifications']),
        'measurement_specs': sorted(all_variant_spec_names['measurement_specs'] - template_spec_names['measurement_specs']),
        'style_specs': sorted(all_variant_spec_names['style_specs'] - template_spec_names['style_specs']),
        'features': sorted(all_variant_spec_names['features'] - template_spec_names['features']),
        'user_guide': sorted(all_variant_spec_names['user_guide'] - template_spec_names['user_guide']),
        'item_details': sorted(all_variant_spec_names['item_details'] - template_spec_names['item_details'])
    }
    
    # Create numbered column names for item_details and user_guide
    additional_item_details_columns = []
    additional_user_guide_columns = []
    
    # For item_details: create columns for each unique name, numbered if duplicates exist
    for name in sorted(all_variant_spec_names['item_details'] - template_spec_names['item_details']):
        max_count = item_details_name_max_count.get(name, 1)
        if max_count > 1:
            # Multiple entries with same name - create numbered columns
            for i in range(1, max_count + 1):
                additional_item_details_columns.append((name, i))
        else:
            # Single entry - no number needed
            additional_item_details_columns.append((name, None))
    
    # For user_guide: create columns for each unique name, numbered if duplicates exist
    for name in sorted(all_variant_spec_names['user_guide'] - template_spec_names['user_guide']):
        max_count = user_guide_name_max_count.get(name, 1)
        if max_count > 1:
            # Multiple entries with same name - create numbered columns
            for i in range(1, max_count + 1):
                additional_user_guide_columns.append((name, i))
        else:
            # Single entry - no number needed
            additional_user_guide_columns.append((name, None))
    
    # Add additional specification columns to child headers
    for spec_name in additional_specs['specifications']:
        child_headers.append(f'Specification: {spec_name}')
    for spec_name in additional_specs['measurement_specs']:
        child_headers.append(f'Measurement Specification: {spec_name}')
    for spec_name in additional_specs['style_specs']:
        child_headers.append(f'Style Specification: {spec_name}')
    for spec_name in additional_specs['features']:
        child_headers.append(f'Feature: {spec_name}')
    # Add user_guide columns (numbered if duplicates)
    for name, number in additional_user_guide_columns:
        if number:
            child_headers.append(f'User Guide: {name} ({number})')
        else:
            child_headers.append(f'User Guide: {name}')
    # Add item_details columns (numbered if duplicates)
    for name, number in additional_item_details_columns:
        if number:
            child_headers.append(f'Item Detail: {name} ({number})')
        else:
            child_headers.append(f'Item Detail: {name}')
    
    # Re-style child header (update headers after adding additional specs)
    for col_num, header in enumerate(child_headers, 1):
        cell = ws_child.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write child variant data
    row_num = 2
    parent_sku = product.sku or ''
    
    for variant in variants:
        col_idx = 1
        ws_child.cell(row=row_num, column=col_idx, value='Child'); col_idx += 1  # Variation Relationship
        ws_child.cell(row=row_num, column=col_idx, value=parent_sku); col_idx += 1  # Parent SKU
        ws_child.cell(row=row_num, column=col_idx, value=variant.color.name if variant.color else ''); col_idx += 1  # Variant Color
        ws_child.cell(row=row_num, column=col_idx, value=variant.size or ''); col_idx += 1  # Variant Size
        ws_child.cell(row=row_num, column=col_idx, value=variant.pattern or ''); col_idx += 1  # Variant Pattern
        ws_child.cell(row=row_num, column=col_idx, value=variant.quality or ''); col_idx += 1  # Variant Quality
        ws_child.cell(row=row_num, column=col_idx, value=variant.title or ''); col_idx += 1  # Variant Title
        ws_child.cell(row=row_num, column=col_idx, value=variant.sku or ''); col_idx += 1  # Variant SKU
        ws_child.cell(row=row_num, column=col_idx, value=float(variant.price) if variant.price else ''); col_idx += 1  # Variant Price
        ws_child.cell(row=row_num, column=col_idx, value=float(variant.old_price) if variant.old_price else ''); col_idx += 1  # Variant Old Price
        ws_child.cell(row=row_num, column=col_idx, value=variant.stock_quantity); col_idx += 1  # Variant Stock Quantity
        ws_child.cell(row=row_num, column=col_idx, value='Yes' if variant.is_in_stock else 'No'); col_idx += 1  # Variant Is In Stock
        ws_child.cell(row=row_num, column=col_idx, value='Yes' if variant.is_active else 'No'); col_idx += 1  # Variant Is Active
        ws_child.cell(row=row_num, column=col_idx, value=variant.image or ''); col_idx += 1  # Variant Image URL
        ws_child.cell(row=row_num, column=col_idx, value=variant.video_url or ''); col_idx += 1  # Variant Video URL
        
        # Multiple variant images (up to 9) - alt text and sort order are auto-generated during import
        variant_images = list(variant.images.all().order_by('sort_order')[:9])
        
        for idx in range(9):
            if idx < len(variant_images):
                img = variant_images[idx]
                ws_child.cell(row=row_num, column=col_idx, value=img.image or '')  # Image URL
                col_idx += 1
            else:
                ws_child.cell(row=row_num, column=col_idx, value='')  # Image URL
                col_idx += 1
        
        # Subcategory boolean columns for variants
        variant_subcategory_ids = set(variant.subcategories.values_list('id', flat=True))
        for subcategory in subcategory_list:
            ws_child.cell(row=row_num, column=col_idx, value='Yes' if subcategory.id in variant_subcategory_ids else 'No')
            col_idx += 1
        
        # Category template specifications (from category templates or defaults)
        # Column headers format: "Specification: {name}", "Style Specification: {name}", etc.
        # Use actual variant data if available, otherwise leave empty
        # col_idx now points to where specifications start
        for spec_name in spec_groups['specifications']:
            spec = variant.specifications.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        for spec_name in spec_groups['measurement_specs']:
            spec = variant.measurement_specs.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        for spec_name in spec_groups['style_specs']:
            spec = variant.style_specs.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        for spec_name in spec_groups['features']:
            spec = variant.features.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        for spec_name in spec_groups['user_guide']:
            spec = variant.user_guide.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        for spec_name in spec_groups['item_details']:
            spec = variant.item_details.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        
        # Additional specifications (not in category templates)
        for spec_name in additional_specs['specifications']:
            spec = variant.specifications.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        for spec_name in additional_specs['measurement_specs']:
            spec = variant.measurement_specs.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        for spec_name in additional_specs['style_specs']:
            spec = variant.style_specs.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        for spec_name in additional_specs['features']:
            spec = variant.features.filter(name=spec_name).first()
            ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
            col_idx += 1
        # Export user_guide entries (matching by name and order)
        user_guide_list = list(variant.user_guide.all().order_by('sort_order', 'id'))
        user_guide_by_name = {}
        for spec in user_guide_list:
            if spec.name not in user_guide_by_name:
                user_guide_by_name[spec.name] = []
            user_guide_by_name[spec.name].append(spec)
        
        for name, number in additional_user_guide_columns:
            if name in user_guide_by_name:
                if number is not None and number <= len(user_guide_by_name[name]):
                    # Get the specific numbered entry
                    spec = user_guide_by_name[name][number - 1]
                    ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
                elif number is None:
                    # Single entry, no number
                    spec = user_guide_by_name[name][0] if user_guide_by_name[name] else None
                    ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
                else:
                    ws_child.cell(row=row_num, column=col_idx, value='')
            else:
                ws_child.cell(row=row_num, column=col_idx, value='')
            col_idx += 1
        
        # Export item_details entries (matching by name and order)
        item_details_list = list(variant.item_details.all().order_by('sort_order', 'id'))
        item_details_by_name = {}
        for spec in item_details_list:
            if spec.name not in item_details_by_name:
                item_details_by_name[spec.name] = []
            item_details_by_name[spec.name].append(spec)
        
        for name, number in additional_item_details_columns:
            if name in item_details_by_name:
                if number is not None and number <= len(item_details_by_name[name]):
                    # Get the specific numbered entry
                    spec = item_details_by_name[name][number - 1]
                    ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
                elif number is None:
                    # Single entry, no number
                    spec = item_details_by_name[name][0] if item_details_by_name[name] else None
                    ws_child.cell(row=row_num, column=col_idx, value=spec.value if spec else '')
                else:
                    ws_child.cell(row=row_num, column=col_idx, value='')
            else:
                ws_child.cell(row=row_num, column=col_idx, value='')
            col_idx += 1
        
        row_num += 1
    
    # Add data validations to both sheets
    # Parent sheet validations
    material_col = get_column_letter(8)
    assembly_col = get_column_letter(13)
    featured_col = get_column_letter(17)
    active_col = get_column_letter(18)
    subcategory_start_col = 22  # After "What is in Box"
    
    if material_names:
        material_dv = DataValidation(type="list", formula1=f'"{",".join(material_names)}"', allow_blank=True)
        ws_parent.add_data_validation(material_dv)
        material_dv.add(f'{material_col}2:{material_col}1000')
    
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws_parent.add_data_validation(yes_no_dv)
    yes_no_dv.add(f'{assembly_col}2:{assembly_col}1000')
    yes_no_dv.add(f'{featured_col}2:{featured_col}1000')
    yes_no_dv.add(f'{active_col}2:{active_col}1000')
    
    # Add Yes/No validation for subcategory boolean columns
    for idx, subcategory in enumerate(subcategory_list):
        subcategory_col = get_column_letter(subcategory_start_col + idx)
        yes_no_dv.add(f'{subcategory_col}2:{subcategory_col}1000')
    
    # Child sheet validations
    color_col = get_column_letter(3)
    stock_status_col = get_column_letter(11)
    variant_subcategory_start_col = 37  # After Variant Image URL (13) + 9 other_image columns (9 columns: 9 images * 1 column each)
    
    if color_names:
        color_dv = DataValidation(type="list", formula1=f'"{",".join(color_names)}"', allow_blank=False)
        ws_child.add_data_validation(color_dv)
        color_dv.add(f'{color_col}2:{color_col}1000')
    
    yes_no_dv_child = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws_child.add_data_validation(yes_no_dv_child)
    yes_no_dv_child.add(f'{stock_status_col}2:{stock_status_col}1000')
    
    # Add Yes/No validation for variant subcategory boolean columns
    for idx, subcategory in enumerate(subcategory_list):
        subcategory_col = get_column_letter(variant_subcategory_start_col + idx)
        yes_no_dv_child.add(f'{subcategory_col}2:{subcategory_col}1000')
    
    # Auto-adjust column widths for both sheets
    for ws in [ws_parent, ws_child]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 15), 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    return wb
