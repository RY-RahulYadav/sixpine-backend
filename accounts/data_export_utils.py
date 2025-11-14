"""Utility functions for exporting user data to Excel"""
import os
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from orders.models import Order, Address
from .models import SavedCard, PaymentPreference


def export_orders_to_excel(user, file_path):
    """Export user orders to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Header row
    headers = [
        'Order ID', 'Order Date', 'Status', 'Payment Status', 'Payment Method',
        'Subtotal', 'Coupon Discount', 'Shipping Cost', 'Platform Fee', 'Tax', 'Total Amount',
        'Shipping Address', 'City', 'State', 'Postal Code', 'Country',
        'Tracking Number', 'Estimated Delivery', 'Delivered At', 'Order Notes'
    ]
    
    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get orders
    orders = Order.objects.filter(user=user).order_by('-created_at')
    
    # Data rows
    for row_num, order in enumerate(orders, 2):
        ws.cell(row=row_num, column=1, value=str(order.order_id))
        ws.cell(row=row_num, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=3, value=order.get_status_display())
        ws.cell(row=row_num, column=4, value=order.get_payment_status_display())
        ws.cell(row=row_num, column=5, value=order.get_payment_method_display() if order.payment_method else 'N/A')
        ws.cell(row=row_num, column=6, value=float(order.subtotal))
        ws.cell(row=row_num, column=7, value=float(order.coupon_discount))
        ws.cell(row=row_num, column=8, value=float(order.shipping_cost))
        ws.cell(row=row_num, column=9, value=float(order.platform_fee))
        ws.cell(row=row_num, column=10, value=float(order.tax_amount))
        ws.cell(row=row_num, column=11, value=float(order.total_amount))
        
        # Shipping address
        if order.shipping_address:
            addr = order.shipping_address
            ws.cell(row=row_num, column=12, value=addr.street_address)
            ws.cell(row=row_num, column=13, value=addr.city)
            ws.cell(row=row_num, column=14, value=addr.state)
            ws.cell(row=row_num, column=15, value=addr.postal_code)
            ws.cell(row=row_num, column=16, value=addr.country)
        
        ws.cell(row=row_num, column=17, value=order.tracking_number or 'N/A')
        ws.cell(row=row_num, column=18, value=order.estimated_delivery.strftime('%Y-%m-%d') if order.estimated_delivery else 'N/A')
        ws.cell(row=row_num, column=19, value=order.delivered_at.strftime('%Y-%m-%d %H:%M:%S') if order.delivered_at else 'N/A')
        ws.cell(row=row_num, column=20, value=order.order_notes or 'N/A')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save file
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    return file_path


def export_addresses_to_excel(user, file_path):
    """Export user addresses to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Addresses"
    
    # Header row
    headers = [
        'Type', 'Full Name', 'Phone', 'Street Address', 'City', 'State',
        'Postal Code', 'Country', 'Is Default', 'Created At', 'Updated At'
    ]
    
    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get addresses
    addresses = Address.objects.filter(user=user).order_by('-created_at')
    
    # Data rows
    for row_num, address in enumerate(addresses, 2):
        ws.cell(row=row_num, column=1, value=address.get_type_display())
        ws.cell(row=row_num, column=2, value=address.full_name)
        ws.cell(row=row_num, column=3, value=address.phone)
        ws.cell(row=row_num, column=4, value=address.street_address)
        ws.cell(row=row_num, column=5, value=address.city)
        ws.cell(row=row_num, column=6, value=address.state)
        ws.cell(row=row_num, column=7, value=address.postal_code)
        ws.cell(row=row_num, column=8, value=address.country)
        ws.cell(row=row_num, column=9, value='Yes' if address.is_default else 'No')
        ws.cell(row=row_num, column=10, value=address.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=11, value=address.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save file
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    return file_path


def export_payment_options_to_excel(user, file_path):
    """Export user payment options to Excel file"""
    wb = Workbook()
    
    # Payment Preferences Sheet
    ws_prefs = wb.active
    ws_prefs.title = "Payment Preferences"
    
    headers_prefs = ['Preferred Method', 'Created At', 'Updated At']
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers_prefs, 1):
        cell = ws_prefs.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get payment preference
    try:
        pref = PaymentPreference.objects.get(user=user)
        ws_prefs.cell(row=2, column=1, value=pref.get_preferred_method_display())
        ws_prefs.cell(row=2, column=2, value=pref.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws_prefs.cell(row=2, column=3, value=pref.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
    except PaymentPreference.DoesNotExist:
        ws_prefs.cell(row=2, column=1, value='No payment preference set')
    
    # Auto-adjust column widths for preferences
    for col in ws_prefs.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_prefs.column_dimensions[col_letter].width = adjusted_width
    
    # Saved Cards Sheet
    ws_cards = wb.create_sheet("Saved Cards")
    
    headers_cards = [
        'Card Last 4', 'Card Network', 'Card Type', 'Card Issuer', 'Nickname',
        'Is Default', 'Created At', 'Updated At'
    ]
    
    for col_num, header in enumerate(headers_cards, 1):
        cell = ws_cards.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get saved cards
    saved_cards = SavedCard.objects.filter(user=user).order_by('-created_at')
    
    for row_num, card in enumerate(saved_cards, 2):
        ws_cards.cell(row=row_num, column=1, value=card.card_last4)
        ws_cards.cell(row=row_num, column=2, value=card.card_network or 'N/A')
        ws_cards.cell(row=row_num, column=3, value=card.card_type or 'N/A')
        ws_cards.cell(row=row_num, column=4, value=card.card_issuer or 'N/A')
        ws_cards.cell(row=row_num, column=5, value=card.nickname or 'N/A')
        ws_cards.cell(row=row_num, column=6, value='Yes' if card.is_default else 'No')
        ws_cards.cell(row=row_num, column=7, value=card.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws_cards.cell(row=row_num, column=8, value=card.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    if saved_cards.count() == 0:
        ws_cards.cell(row=2, column=1, value='No saved cards')
    
    # Auto-adjust column widths for cards
    for col in ws_cards.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_cards.column_dimensions[col_letter].width = adjusted_width
    
    # Save file
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    return file_path

